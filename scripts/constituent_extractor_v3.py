
"""
Constituent Contact Extractor v3 (Final)
=========================================
Extracts missing contact info from Salesforce case text fields with:
- LLM disambiguation for ambiguous cases
- Improved address extraction (handles St. George St, lowercase, etc.)
- Improved name extraction (signatures, "my name is", etc.)
- Email chain detection
- Duplicate grouping
- Constituent vs complaint address detection

Usage:
    python constituent_extractor_v3.py input.csv output.xlsx --api-key YOUR_KEY

"""
import os
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()
import argparse
import json
import re
import time
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False


# config

TEXT_FIELDS = ['Subject', 'Description', 'Case Notes', 'Case Comments']

STAFF_EMAILS = {
    'bob@diannesaxe.ca',
    'sydney@diannesaxe.ca',
    'christian.cullis@toronto.ca',
    'councillor_saxe@toronto.ca',
    'dianne.saxe@toronto.ca',
    'dsaxe@envirolaw.com',
    'saxeforcouncil@gmail.com',
    'anne@diannesaxe.ca',
    'william.hopkins2@toronto.ca',
}

STAFF_EMAIL_PATTERNS = ['@diannesaxe.ca', '@toronto.ca', '@envirolaw.com']

VALID_AREA_CODES = {'416', '647', '437', '905', '289', '365', '519', '226', '548', '613', '343', '705', '249', '807'}


#regex patterns

EMAIL_PATTERN = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', re.IGNORECASE)

PHONE_10_PATTERN = re.compile(r'\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4}')
PHONE_7_PATTERN = re.compile(r'(?<!\d)\d{3}[\s.\-]?\d{4}(?!\d)')

# Toronto postal code
POSTAL_PATTERN = re.compile(r'\b[Mm]\d[A-Za-z][\s]?\d[A-Za-z]\d\b')

# Street suffixes
STREET_SUFFIXES = r'(?:St(?:reet)?|Ave(?:nue)?|Rd|Road|Dr(?:ive)?|Blvd|Boulevard|Cres(?:cent)?|Ct|Court|Pl(?:ace)?|Terr(?:ace)?|Lane|Ln|Way|Circle|Cir|Trail|Trl|Park|Pk|Gardens?|Gate|Grove|Heights?|Hts|Hill|Square|Sq|Row|Walk|Mews)'

# Address patterns - multiple approaches to catch more
ADDRESS_PATTERNS = [
    # Standard: 123 Main St
    re.compile(
        rf'(\d+[A-Za-z]?(?:\s*[-–]\s*\d+[A-Za-z]?)?)\s+'
        rf'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s+'
        rf'{STREET_SUFFIXES}\.?'
        rf'(?:\s*(?:East|West|North|South|E\.?|W\.?|N\.?|S\.?))?',
        re.IGNORECASE
    ),
    # ALL CAPS: 690 MANNING AVE
    re.compile(
        rf'(\d+[A-Za-z]?)\s+([A-Z]{{2,}}(?:\s+[A-Z]{{2,}})?)\s+{STREET_SUFFIXES}',
        re.IGNORECASE
    ),
    # St. Name Street (like St. George St, St. Clair Ave)
    re.compile(
        rf'(\d+[A-Za-z]?)\s+(St\.?\s+[A-Z][a-z]+)\s+(?:Street|St\.?|Ave(?:nue)?|Rd|Road)',
        re.IGNORECASE
    ),
    # Lowercase: 120 macpherson ave
    re.compile(
        rf'(\d+[A-Za-z]?)\s+([a-z]+(?:\s+[a-z]+)?)\s+{STREET_SUFFIXES}',
        re.IGNORECASE
    ),
    # With unit: Unit 5, 123 Main St or Apt 4, 123 Main St
    re.compile(
        rf'(?:Unit|Apt|Suite|#)\s*\d+[A-Za-z]?\s*,?\s*(\d+)\s+([A-Za-z]+(?:\s+[A-Za-z]+)?)\s+{STREET_SUFFIXES}',
        re.IGNORECASE
    ),
]

# Email chain indicators
EMAIL_CHAIN_PATTERNS = [
    re.compile(r'^From:\s*', re.MULTILINE | re.IGNORECASE),
    re.compile(r'^Sent:\s*', re.MULTILINE | re.IGNORECASE),
    re.compile(r'[-]{3,}.*forwarded.*[-]{3,}', re.IGNORECASE),
    re.compile(r'[-]{3,}.*original message.*[-]{3,}', re.IGNORECASE),
    re.compile(r'^>+\s*', re.MULTILINE),
    re.compile(r'wrote:$', re.MULTILINE | re.IGNORECASE),
]

# Constituent address indicators
CONSTITUENT_ADDRESS_PHRASES = [
    r'i\s+live\s+at',
    r'i\s+reside\s+at',
    r'i\s+am\s+(?:a\s+)?resident\s+(?:of|at)',
    r'my\s+(?:home\s+)?address\s+is',
    r'we\s+live\s+at',
    r'our\s+(?:home\s+)?address',
    r'my\s+property\s+at',
    r'i\s+own\s+(?:a\s+)?(?:home|house|property)\s+at',
    r'we\s+own\s+(?:a\s+)?(?:home|house|property)\s+at',
    r'my\s+house\s+(?:is\s+)?at',
    r'i\s+am\s+(?:the\s+)?owner\s+of',
]

# Name extraction patterns
NAME_PATTERNS = [
    # "My name is John Smith"
    (re.compile(r'my\s+name\s+is\s+([A-Z][a-z]+(?:\s+[A-Z][a-z\-]+)+)', re.IGNORECASE), 'intro'),
    # "I am John Smith" or "I'm John Smith" (at start of sentence)
    (re.compile(r"(?:^|[.!?]\s+)I'?m\s+([A-Z][a-z]+\s+[A-Z][a-z\-]+)(?=[\s,.])", re.MULTILINE), 'intro'),
    # "This is John Smith" (calling/writing)
    (re.compile(r'this\s+is\s+([A-Z][a-z]+\s+[A-Z][a-z\-]+)(?:\s+(?:calling|writing|from))', re.IGNORECASE), 'intro'),
    # Sign-off: "Regards,\n\nJohn Smith" - stop at newline or common trailing words
    (re.compile(r'(?:Yours\s+truly|Sincerely|Best\s+regards?|Thanks|Thank\s+you|Regards|Cheers|Best|Warmly)[,.\s]*\r?\n+\s*([A-Z][a-z]+(?:\s+[A-Z][a-z\-]+)+)(?=\s*(?:\r?\n|$|[,\r]))', re.IGNORECASE), 'signoff'),
    # Name on its own line followed by email/phone on next line
    (re.compile(r'^\s*([A-Z][a-z]+\s+[A-Z][a-z\-]+)\s*$\n.*?(?:@|\(\d{3}\)|\d{3}[\-.])', re.MULTILINE), 'contact_block'),
    # Name at very end of text (last non-empty line)
    (re.compile(r'\r?\n\s*([A-Z][a-z]+(?:\s+[A-Z][a-z\-]+)+)\s*$'), 'end'),
]

# False positive names to filter
FALSE_POSITIVE_NAMES = {
    'dear councillor', 'dear dianne', 'hello councillor', 'hi councillor',
    'dear ms', 'dear mr', 'dear dr', 'hello dianne', 'hi dianne',
    'councillor saxe', 'dianne saxe', 'dr saxe', 'ms saxe',
    'city of', 'university of', 'town of', 'province of',
    'best regards', 'kind regards', 'warm regards', 'on behalf',
    'sent from', 'get outlook',
}


# utility functions

def safe_str(value) -> str:
    """Convert value to string, handling NaN/None."""
    if pd.isna(value) or value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() == 'nan' else s


def is_staff_email(email: str) -> bool:
    """Check if email belongs to staff."""
    if not email:
        return False
    email_lower = email.lower()
    if email_lower in STAFF_EMAILS:
        return True
    return any(pattern in email_lower for pattern in STAFF_EMAIL_PATTERNS)


def validate_postal(postal: str) -> Tuple[bool, str]:
    """Validate Toronto postal code."""
    if not postal:
        return False, ""
    normalized = postal.upper().replace(" ", "")
    if not normalized.startswith('M') or len(normalized) != 6:
        return False, normalized
    if not re.match(r'^M\d[A-Z]\d[A-Z]\d$', normalized):
        return False, normalized
    return True, f"{normalized[:3]} {normalized[3:]}"


def validate_phone(phone: str) -> Tuple[str, str]:
    """Validate phone number. Returns (status, normalized)."""
    if not phone:
        return 'invalid', ""
    digits = re.sub(r'\D', '', phone)
    if len(digits) == 10:
        return 'valid', f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    elif len(digits) == 7:
        return 'partial', f"{digits[:3]}-{digits[3:]}"
    return 'invalid', phone


def detect_email_chain(text: str) -> bool:
    """Detect if text contains forwarded/chain email indicators."""
    if not text:
        return False
    return any(pattern.search(text) for pattern in EMAIL_CHAIN_PATTERNS)


def is_constituent_address_context(text: str, address: str) -> bool:
    """Check if address appears in context suggesting it's the constituent's."""
    if not text or not address:
        return False
    text_lower = text.lower()
    addr_lower = address.lower()
    
    # Find position of address
    addr_pos = text_lower.find(addr_lower[:20])  # First 20 chars
    if addr_pos == -1:
        return False
    
    # Check 150 chars before for constituent phrases
    context = text_lower[max(0, addr_pos - 150):addr_pos]
    return any(re.search(phrase, context) for phrase in CONSTITUENT_ADDRESS_PHRASES)


# extraction functions

def extract_addresses(text: str) -> List[str]:
    """Extract street addresses using multiple patterns."""
    if not text:
        return []
    
    addresses = set()
    
    for pattern in ADDRESS_PATTERNS:
        for match in pattern.finditer(text):
            # Use the full match, not just the groups
            addr = match.group(0).strip()
            addr = re.sub(r'\s+', ' ', addr)  # Normalize whitespace
            addresses.add(addr)
    
    # Filter false positives
    false_positives = [
        'minute', 'minutes', 'hour', 'hours', 'day', 'days', 'week', 'weeks',
        'year', 'years', 'month', 'months', 'meter', 'meters', 'dogs', 'cats',
        'cars', 'people', 'times', 'units', 'feet', 'inches', 'miles', 'block',
        'children', 'students', 'residents', 'members', 'guests'
    ]
    
    filtered = []
    for addr in addresses:
        addr_lower = addr.lower()
        if not any(word in addr_lower for word in false_positives):
            # Minimum length of 6 to allow "6 Bay St" type addresses
            if len(addr) >= 6:
                filtered.append(addr)
    
    return list(set(filtered))


def extract_names(text: str) -> List[Tuple[str, str]]:
    """Extract names using multiple patterns. Returns (name, pattern_type) tuples."""
    if not text:
        return []
    
    names = []
    
    for pattern, pattern_type in NAME_PATTERNS:
        for match in pattern.finditer(text):
            name = match.group(1).strip()
            # Clean up the name
            name = clean_extracted_name(name)
            if name:
                names.append((name, pattern_type))
    
    # Deduplicate while preserving order
    seen = set()
    unique_names = []
    for name, ptype in names:
        if name.lower() not in seen:
            seen.add(name.lower())
            unique_names.append((name, ptype))
    
    return unique_names


def clean_extracted_name(name: str) -> Optional[str]:
    """Clean and validate an extracted name."""
    if not name:
        return None
    
    # Remove common trailing junk
    name = re.sub(r'\s*[\r\n].*$', '', name)  # Remove everything after newline
    name = re.sub(r'\s*\d+.*$', '', name)  # Remove trailing numbers (dates, addresses)
    
    # Remove trailing phrases
    trailing_phrases = [
        r'\s+on behalf.*$',
        r'\s+sent from.*$',
        r'\s+get outlook.*$',
        r'\s+\(.*\)$',  # Remove parenthetical at end
    ]
    for phrase in trailing_phrases:
        name = re.sub(phrase, '', name, flags=re.IGNORECASE)
    
    name = name.strip(' ,.')
    
    # Filter false positives
    if name.lower() in FALSE_POSITIVE_NAMES:
        return None
    
    # Must have at least 2 words
    words = name.split()
    if len(words) < 2:
        return None
    
    # Each word should be at least 2 chars
    if not all(len(w) >= 2 for w in words):
        return None
    
    # Should look like a name (capitalized words)
    if not all(w[0].isupper() for w in words):
        return None
    
    # Filter names that are too long (likely sentences)
    if len(words) > 5:
        return None
    
    return name


def extract_emails(text: str) -> List[str]:
    """Extract non-staff emails."""
    if not text:
        return []
    all_emails = EMAIL_PATTERN.findall(text)
    return list(set(e for e in all_emails if not is_staff_email(e)))


def extract_phones(text: str) -> List[Tuple[str, str]]:
    """Extract and validate phone numbers. Returns (status, normalized) tuples."""
    if not text:
        return []
    
    phones = []
    seen = set()
    
    for match in PHONE_10_PATTERN.finditer(text):
        status, normalized = validate_phone(match.group())
        if status != 'invalid' and normalized not in seen:
            phones.append((status, normalized))
            seen.add(normalized)
    
    for match in PHONE_7_PATTERN.finditer(text):
        status, normalized = validate_phone(match.group())
        if status == 'partial' and normalized not in seen:
            phones.append((status, normalized))
            seen.add(normalized)
    
    return phones


def extract_postals(text: str) -> List[Tuple[bool, str]]:
    """Extract and validate postal codes. Returns (is_valid, normalized) tuples."""
    if not text:
        return []
    
    postals = []
    seen = set()
    
    for match in POSTAL_PATTERN.finditer(text):
        is_valid, normalized = validate_postal(match.group())
        if normalized and normalized not in seen:
            postals.append((is_valid, normalized))
            seen.add(normalized)
    
    return postals


# llm disambiguation

LLM_PROMPT = """You are extracting constituent contact information from a Toronto city councillor's case management system.

CASE TEXT:
{text}

REGEX EXTRACTED THESE CANDIDATES:
- Emails: {emails}
- Phones: {phones}
- Addresses: {addresses}

YOUR TASK:
Identify which contact info belongs to THE CONSTITUENT (the person writing to ask for help from the councillor).

DO NOT select:
- Staff emails (anything @toronto.ca or @diannesaxe.ca)
- People being complained about
- Senders of forwarded emails (unless they ARE the constituent)
- Organizations that are CC'd
- Addresses that are complaint locations, not where the constituent lives

IMPORTANT DISTINCTIONS:
- "I live at 400 Markham St" = constituent address
- "The property at 400 Markham is causing problems" = complaint address
- In forwarded emails, the person who forwarded TO the councillor is usually the constituent

Return ONLY valid JSON (no markdown, no explanation):
{{
  "constituent_email": "email@example.com or null",
  "constituent_phone": "(416) 555-1234 or null",
  "constituent_address": "123 Main St or null",
  "complaint_address": "456 Other St or null if same as constituent or not mentioned",
  "confidence": "high/medium/low",
  "reasoning": "One sentence explanation"
}}"""


def call_llm(text: str, emails: List[str], phones: List[Tuple[str, str]], 
             addresses: List[str], client) -> Optional[Dict]:
    """Call Claude to disambiguate contact info."""
    try:
        prompt = LLM_PROMPT.format(
            text=text[:4000],  # Truncate for token limits
            emails=', '.join(emails) if emails else 'None found',
            phones=', '.join(f"{s}:{p}" for s, p in phones) if phones else 'None found',
            addresses=', '.join(addresses) if addresses else 'None found',
        )
        
        response = client.messages.create(
            model="claude-3-5-haiku-20241022",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}]
        )
        
        # Parse JSON from response
        response_text = response.content[0].text.strip()
        # Try to extract JSON even if there's extra text
        json_match = re.search(r'\{[^{}]*\}', response_text, re.DOTALL)
        if json_match:
            return json.loads(json_match.group())
        return None
        
    except Exception as e:
        return {'error': str(e)}


# main extraction
@dataclass
class ExtractionResult:
    """Complete extraction results for a case."""
    # Salesforce data
    sf_name: str = ""
    sf_email: str = ""
    sf_phone: str = ""
    sf_has_contact_id: bool = False
    sf_is_contaminated: bool = False
    
    # Extracted data
    extracted_emails: List[str] = field(default_factory=list)
    extracted_phones: List[Tuple[str, str]] = field(default_factory=list)
    extracted_addresses: List[Tuple[str, bool]] = field(default_factory=list)  # (addr, is_constituent)
    extracted_postals: List[Tuple[bool, str]] = field(default_factory=list)  # (is_valid, postal)
    extracted_names: List[Tuple[str, str]] = field(default_factory=list)  # (name, pattern_type)
    
    # LLM results (if used)
    llm_email: Optional[str] = None
    llm_phone: Optional[str] = None
    llm_address: Optional[str] = None
    llm_complaint_address: Optional[str] = None
    llm_confidence: Optional[str] = None
    llm_reasoning: str = ""
    llm_used: bool = False
    
    # Flags
    is_email_chain: bool = False
    has_multiple_emails: bool = False
    has_multiple_addresses: bool = False
    matched_contact_id: Optional[str] = None
    
    # Computed fields
    missing_fields: List[str] = field(default_factory=list)
    extractable_fields: List[str] = field(default_factory=list)
    action: str = ""
    confidence: str = ""
    flags: List[str] = field(default_factory=list)


def needs_llm(result: ExtractionResult) -> bool:
    """Determine if this case needs LLM disambiguation."""
    # Multiple emails - definitely need to pick the right one
    if len(result.extracted_emails) > 1:
        return True
    
    # Multiple addresses AND none are flagged as constituent
    if len(result.extracted_addresses) > 1:
        has_constituent = any(is_const for _, is_const in result.extracted_addresses)
        if not has_constituent:
            return True
    
    # Email chain with multiple addresses - forwarded content may be confusing
    if result.is_email_chain and len(result.extracted_addresses) > 1:
        return True
    
    # Email chain with extracted email but no clear single email
    if result.is_email_chain and result.extracted_emails and not result.sf_email:
        return True
    
    return False


def extract_from_case(row: pd.Series, email_to_contact: Dict[str, str], 
                      client=None, use_llm: bool = False) -> ExtractionResult:
    """Extract all contact info from a case."""
    result = ExtractionResult()
    
    # === Salesforce data ===
    result.sf_name = safe_str(row.get('Contact Name'))
    result.sf_email = safe_str(row.get('Contact: Email'))
    web_email = safe_str(row.get('Web Email'))
    result.sf_phone = safe_str(row.get('Contact: Phone'))
    
    contact_id = safe_str(row.get('Contact ID'))
    result.sf_has_contact_id = bool(contact_id and contact_id != '000000000000000')
    
    result.sf_is_contaminated = is_staff_email(result.sf_email) or is_staff_email(web_email)
    
    if not result.sf_email and web_email and not is_staff_email(web_email):
        result.sf_email = web_email
    
    # === Combine text fields ===
    text_parts = []
    for field_name in TEXT_FIELDS:
        value = row.get(field_name)
        if pd.notna(value) and str(value).strip():
            text_parts.append(str(value))
    combined_text = "\n".join(text_parts)
    
    if not combined_text.strip():
        result.action = "NO_TEXT"
        return result
    
    # === Detect email chain ===
    result.is_email_chain = detect_email_chain(combined_text)
    if result.is_email_chain:
        result.flags.append("EMAIL_CHAIN")
    
    # === Extract all candidates ===
    result.extracted_emails = extract_emails(combined_text)
    result.extracted_phones = extract_phones(combined_text)
    
    raw_addresses = extract_addresses(combined_text)
    result.extracted_addresses = [
        (addr, is_constituent_address_context(combined_text, addr))
        for addr in raw_addresses
    ]
    
    result.extracted_postals = extract_postals(combined_text)
    result.extracted_names = extract_names(combined_text)
    
    # === Set flags ===
    if len(result.extracted_emails) > 1:
        result.has_multiple_emails = True
        result.flags.append("MULTIPLE_EMAILS")
    
    if len(result.extracted_addresses) > 1:
        result.has_multiple_addresses = True
        result.flags.append("MULTIPLE_ADDRESSES")
    
    # === LLM disambiguation if needed ===
    if use_llm and client and needs_llm(result):
        llm_result = call_llm(
            combined_text,
            result.extracted_emails,
            result.extracted_phones,
            [addr for addr, _ in result.extracted_addresses],
            client
        )
        
        if llm_result and 'error' not in llm_result:
            result.llm_used = True
            result.llm_email = llm_result.get('constituent_email')
            result.llm_phone = llm_result.get('constituent_phone')
            result.llm_address = llm_result.get('constituent_address')
            result.llm_complaint_address = llm_result.get('complaint_address')
            result.llm_confidence = llm_result.get('confidence')
            result.llm_reasoning = llm_result.get('reasoning', '')
            result.flags.append("LLM_USED")
        elif llm_result and 'error' in llm_result:
            result.flags.append(f"LLM_ERROR:{llm_result['error'][:50]}")
    
    # === Match unlinked to existing contacts ===
    if not result.sf_has_contact_id:
        emails_to_check = [result.llm_email] if result.llm_email else result.extracted_emails
        for email in emails_to_check:
            if email and email.lower() in email_to_contact:
                result.matched_contact_id = email_to_contact[email.lower()]
                result.flags.append(f"MATCHES:{result.matched_contact_id[:15]}")
                break
    
    # === Determine missing/extractable fields ===
    if not result.sf_name:
        result.missing_fields.append("name")
    if not result.sf_email:
        result.missing_fields.append("email")
    if not result.sf_phone:
        result.missing_fields.append("phone")
    
    # What can we extract?
    best_email = result.llm_email or (result.extracted_emails[0] if len(result.extracted_emails) == 1 else None)
    if best_email and not result.sf_email:
        result.extractable_fields.append("email")
    
    valid_phones = [p for s, p in result.extracted_phones if s == 'valid']
    partial_phones = [p for s, p in result.extracted_phones if s == 'partial']
    best_phone = result.llm_phone or (valid_phones[0] if valid_phones else None)
    
    if best_phone and not result.sf_phone:
        result.extractable_fields.append("phone")
    elif partial_phones and not result.sf_phone:
        result.extractable_fields.append("phone (partial)")
    
    best_address = result.llm_address
    if not best_address and result.extracted_addresses:
        # Prefer constituent-flagged addresses
        constituent_addrs = [a for a, is_c in result.extracted_addresses if is_c]
        if constituent_addrs:
            best_address = constituent_addrs[0]
        elif len(result.extracted_addresses) == 1:
            best_address = result.extracted_addresses[0][0]
    
    if best_address:
        result.extractable_fields.append("address")
    
    valid_postals = [p for v, p in result.extracted_postals if v]
    if valid_postals:
        result.extractable_fields.append("postal")
    elif result.extracted_postals:
        result.extractable_fields.append("postal (invalid)")
    
    best_name = result.extracted_names[0][0] if result.extracted_names else None
    if best_name and not result.sf_name:
        result.extractable_fields.append("name")
    
    # === Determine action ===
    result.action, result.confidence = determine_action(result)
    
    return result


def determine_action(result: ExtractionResult) -> Tuple[str, str]:
    """Determine recommended action and confidence."""
    
    if result.sf_is_contaminated:
        has_recovery = bool(
            result.llm_email or 
            (result.extracted_emails and len(result.extracted_emails) == 1)
        )
        if has_recovery:
            return "FIX_CONTAMINATION_CAN_RECOVER", "medium"
        return "FIX_CONTAMINATION_MANUAL", "low"
    
    if result.sf_has_contact_id:
        if result.extractable_fields:
            conf = "high" if not result.is_email_chain else "medium"
            return "CAN_ENRICH", conf
        return "COMPLETE", "high"
    
    # No Contact ID
    if result.matched_contact_id:
        return "LINK_TO_EXISTING", "high"
    
    has_email = bool(
        result.llm_email or 
        (result.extracted_emails and len(result.extracted_emails) == 1)
    )
    
    if has_email:
        if result.is_email_chain or result.has_multiple_emails:
            if result.llm_used and result.llm_confidence in ['high', 'medium']:
                return "CREATE_CONTACT", "medium"
            return "CREATE_CONTACT_REVIEW", "low"
        return "CREATE_CONTACT", "medium"
    
    if result.extracted_phones:
        return "HAS_PHONE_ONLY", "low"
    
    return "NO_CONTACT_INFO", "none"


# duplicate detection

def find_duplicates(df: pd.DataFrame, results: List[ExtractionResult]) -> Dict[str, List[int]]:
    """Find unlinked cases with same email."""
    email_to_rows = defaultdict(list)
    
    for idx, result in enumerate(results):
        if not result.sf_has_contact_id:
            email = result.llm_email or (result.extracted_emails[0] if result.extracted_emails else None)
            if email:
                email_to_rows[email.lower()].append(idx)
    
    return {email: rows for email, rows in email_to_rows.items() if len(rows) > 1}


# main processing

def process_cases(input_file: str, output_file: str, api_key: Optional[str] = None,
                  use_llm: bool = True, llm_limit: int = 500) -> dict:
    """Process all cases."""
    
    print(f"Loading {input_file}...")
    df = pd.read_csv(input_file, encoding='utf-8-sig')
    print(f"Loaded {len(df)} cases")
    
    # Initialize LLM client
    client = None
    if use_llm and api_key and HAS_ANTHROPIC:
        client = anthropic.Anthropic(api_key=api_key)
        print("LLM disambiguation enabled")
    else:
        print("LLM disambiguation disabled")
        use_llm = False
    
    # Build email index
    print("Building email index...")
    email_to_contact = {}
    for _, row in df.iterrows():
        contact_id = safe_str(row.get('Contact ID'))
        if contact_id and contact_id != '000000000000000':
            for col in ['Contact: Email', 'Web Email']:
                email = safe_str(row.get(col)).lower()
                if email and not is_staff_email(email):
                    email_to_contact[email] = contact_id
    print(f"Found {len(email_to_contact)} unique contact emails")
    
    # Process cases
    print("Extracting contact info...")
    results = []
    llm_calls = 0
    
    for idx, row in df.iterrows():
        if idx % 500 == 0:
            print(f"  Processing case {idx}/{len(df)}... (LLM calls: {llm_calls})")
        
        # Check if we should use LLM for this case
        should_use_llm = use_llm and llm_calls < llm_limit
        
        result = extract_from_case(row, email_to_contact, client if should_use_llm else None, should_use_llm)
        
        if result.llm_used:
            llm_calls += 1
            time.sleep(0.1)  # Rate limiting
        
        results.append(result)
    
    # Find duplicates
    print("Finding duplicate groups...")
    duplicates = find_duplicates(df, results)
    print(f"Found {len(duplicates)} emails with multiple unlinked cases")
    
    # Mark duplicates
    for email, indices in duplicates.items():
        for idx in indices:
            results[idx].flags.append(f"DUP_GROUP:{email[:20]}({len(indices)})")
    
    # Build output dataframe
    print("Building output...")
    
    df['Action'] = [r.action for r in results]
    df['Confidence'] = [r.confidence for r in results]
    df['Missing_Fields'] = [', '.join(r.missing_fields) or 'None' for r in results]
    df['Extractable_Fields'] = [', '.join(r.extractable_fields) or 'None' for r in results]
    
    df['SF_Has_Name'] = ['Yes' if r.sf_name else 'No' for r in results]
    df['SF_Has_Email'] = ['Yes' if r.sf_email and not r.sf_is_contaminated else 'No' for r in results]
    df['SF_Has_Phone'] = ['Yes' if r.sf_phone else 'No' for r in results]
    df['SF_Is_Contaminated'] = ['Yes' if r.sf_is_contaminated else 'No' for r in results]
    
    df['Extracted_Emails'] = ['; '.join(r.extracted_emails) for r in results]
    df['Extracted_Phones'] = ['; '.join(f"{s}:{p}" for s, p in r.extracted_phones) for r in results]
    df['Extracted_Addresses'] = ['; '.join(f"{'[CONST] ' if c else ''}{a}" for a, c in r.extracted_addresses) for r in results]
    df['Extracted_Postals'] = ['; '.join(f"{'✓' if v else '✗'}{p}" for v, p in r.extracted_postals) for r in results]
    df['Extracted_Names'] = ['; '.join(f"{n} ({t})" for n, t in r.extracted_names) for r in results]
    
    df['LLM_Email'] = [r.llm_email or '' for r in results]
    df['LLM_Phone'] = [r.llm_phone or '' for r in results]
    df['LLM_Address'] = [r.llm_address or '' for r in results]
    df['LLM_Complaint_Addr'] = [r.llm_complaint_address or '' for r in results]
    df['LLM_Confidence'] = [r.llm_confidence or '' for r in results]
    df['LLM_Reasoning'] = [r.llm_reasoning for r in results]
    
    df['Matched_Contact_ID'] = [r.matched_contact_id or '' for r in results]
    df['Flags'] = ['; '.join(r.flags) for r in results]
    
    # Stats
    stats = {
        'total': len(df),
        'actions': df['Action'].value_counts().to_dict(),
        'llm_calls': llm_calls,
        'with_contact_id': sum(1 for r in results if r.sf_has_contact_id),
        'contaminated': sum(1 for r in results if r.sf_is_contaminated),
        'duplicate_groups': len(duplicates),
        'cases_in_duplicates': sum(len(rows) for rows in duplicates.values()),
    }
    
    stats['extractable'] = {
        'email': sum(1 for r in results if 'email' in r.extractable_fields),
        'phone': sum(1 for r in results if 'phone' in r.extractable_fields or 'phone (partial)' in r.extractable_fields),
        'address': sum(1 for r in results if 'address' in r.extractable_fields),
        'postal': sum(1 for r in results if 'postal' in r.extractable_fields),
        'name': sum(1 for r in results if 'name' in r.extractable_fields),
    }
    
    # Write Excel
    print(f"Writing to {output_file}...")
    create_excel_output(df, output_file, stats, duplicates)
    
    # Print summary
    print("\n" + "="*70)
    print("PROCESSING COMPLETE")
    print("="*70)
    print(f"\nTotal cases: {stats['total']}")
    print(f"LLM calls made: {stats['llm_calls']}")
    print(f"With Contact ID: {stats['with_contact_id']}")
    print(f"Staff contaminated: {stats['contaminated']}")
    
    print(f"\n=== ACTIONS ===")
    for action, count in sorted(stats['actions'].items(), key=lambda x: -x[1]):
        print(f"  {action}: {count}")
    
    print(f"\n=== EXTRACTABLE FIELDS ===")
    for field, count in stats['extractable'].items():
        print(f"  {field}: {count}")
    
    print(f"\n=== DUPLICATES ===")
    print(f"  {stats['duplicate_groups']} email addresses in multiple unlinked cases")
    print(f"  {stats['cases_in_duplicates']} total cases in duplicate groups")
    
    print(f"\nOutput: {output_file}")
    return stats


def create_excel_output(df: pd.DataFrame, output_file: str, stats: dict, duplicates: dict):
    """Create formatted Excel output."""
    
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions['A'].width = 45
    ws_summary.column_dimensions['B'].width = 15
    
    row = 1
    ws_summary[f'A{row}'] = "CASE ANALYSIS SUMMARY"
    ws_summary[f'A{row}'].font = Font(bold=True, size=16)
    row += 2
    
    ws_summary[f'A{row}'] = "Total Cases"
    ws_summary[f'B{row}'] = stats['total']
    row += 1
    ws_summary[f'A{row}'] = "LLM Calls Made"
    ws_summary[f'B{row}'] = stats['llm_calls']
    row += 1
    ws_summary[f'A{row}'] = "With Contact ID"
    ws_summary[f'B{row}'] = stats['with_contact_id']
    row += 1
    ws_summary[f'A{row}'] = "Staff Email Contaminated"
    ws_summary[f'B{row}'] = stats['contaminated']
    row += 2
    
    ws_summary[f'A{row}'] = "ACTIONS"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    action_colors = {
        'COMPLETE': '90EE90',
        'CAN_ENRICH': 'FFFACD',
        'LINK_TO_EXISTING': 'ADD8E6',
        'CREATE_CONTACT': '87CEEB',
        'CREATE_CONTACT_REVIEW': 'B0E0E6',
        'FIX_CONTAMINATION_CAN_RECOVER': 'FFB6C1',
        'FIX_CONTAMINATION_MANUAL': 'FF9999',
        'HAS_PHONE_ONLY': 'DDA0DD',
        'NO_CONTACT_INFO': 'D3D3D3',
        'NO_TEXT': 'C0C0C0',
    }
    
    for action, count in sorted(stats['actions'].items(), key=lambda x: -x[1]):
        ws_summary[f'A{row}'] = action
        ws_summary[f'B{row}'] = count
        if action in action_colors:
            ws_summary[f'A{row}'].fill = PatternFill('solid', fgColor=action_colors[action])
        row += 1
    
    row += 1
    ws_summary[f'A{row}'] = "EXTRACTABLE FIELDS"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    for field, count in stats['extractable'].items():
        ws_summary[f'A{row}'] = f"Could add {field}"
        ws_summary[f'B{row}'] = count
        row += 1
    
    row += 1
    ws_summary[f'A{row}'] = "DUPLICATES"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    ws_summary[f'A{row}'] = "Emails in multiple unlinked cases"
    ws_summary[f'B{row}'] = stats['duplicate_groups']
    row += 1
    ws_summary[f'A{row}'] = "Total cases in duplicate groups"
    ws_summary[f'B{row}'] = stats['cases_in_duplicates']
    
    # All Cases sheet
    ws_all = wb.create_sheet("All Cases")
    for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor='DDDDDD')
    
    # Color by action
    action_col = list(df.columns).index('Action') + 1
    for r_idx in range(2, len(df) + 2):
        action = ws_all.cell(row=r_idx, column=action_col).value
        if action in action_colors:
            for c_idx in range(1, min(len(df.columns) + 1, 40)):
                ws_all.cell(row=r_idx, column=c_idx).fill = PatternFill('solid', fgColor=action_colors[action])
    
    # Filtered sheets
    sheets = [
        ('CAN_ENRICH', 'Enrich'),
        ('LINK_TO_EXISTING', 'Link'),
        ('CREATE_CONTACT', 'Create'),
        ('CREATE_CONTACT_REVIEW', 'Create-Review'),
        ('FIX_CONTAMINATION_CAN_RECOVER', 'Fix-Recover'),
        ('FIX_CONTAMINATION_MANUAL', 'Fix-Manual'),
    ]
    
    for action, name in sheets:
        action_df = df[df['Action'] == action]
        if len(action_df) > 0:
            ws = wb.create_sheet(name)
            for r_idx, row_data in enumerate(dataframe_to_rows(action_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
    
    # Duplicates sheet
    if duplicates:
        ws_dup = wb.create_sheet("Duplicates")
        ws_dup['A1'] = "Email"
        ws_dup['B1'] = "Count"
        ws_dup['C1'] = "Case Numbers"
        ws_dup['A1'].font = Font(bold=True)
        ws_dup['B1'].font = Font(bold=True)
        ws_dup['C1'].font = Font(bold=True)
        
        row = 2
        for email, indices in sorted(duplicates.items(), key=lambda x: -len(x[1])):
            case_nums = [str(df.iloc[i]['Case Number']) for i in indices]
            ws_dup[f'A{row}'] = email
            ws_dup[f'B{row}'] = len(indices)
            ws_dup[f'C{row}'] = ', '.join(case_nums)
            row += 1
        
        ws_dup.column_dimensions['A'].width = 40
        ws_dup.column_dimensions['C'].width = 80
    
    wb.save(output_file)


# CLI
def main():
    parser = argparse.ArgumentParser(description='Extract constituent contact info from Salesforce cases')
    parser.add_argument('input', help='Input CSV file')
    parser.add_argument('output', help='Output Excel file')
    parser.add_argument('--api-key', help='Anthropic API key')
    parser.add_argument('--no-llm', action='store_true', help='Disable LLM')
    parser.add_argument('--llm-limit', type=int, default=500, help='Max LLM calls')
    
    args = parser.parse_args()
    
    api_key = args.api_key
    if not api_key:
        import os
        api_key = os.environ.get('ANTHROPIC_API_KEY')
    
    process_cases(
        args.input,
        args.output,
        api_key=api_key,
        use_llm=not args.no_llm,
        llm_limit=args.llm_limit,
    )


if __name__ == '__main__':
    main()
