
"""
Outlook to Salesforce Matcher
=
Compares exported Outlook emails against Salesforce cases to identify
emails that were never entered into the system.

Usage:
    python outlook_matcher.py salesforce_cases.csv outlook_export.csv output.xlsx

Input formats:
    - Salesforce: Standard case export with Contact: Email, Web Email columns
    - Outlook: Export from Access/VBA with SenderEmail, Subject, ReceivedDate columns
      (Column names are flexible - script will try to detect them)
"""

import argparse
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Set, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


#config

# Staff/system emails to exclude from "missed" list
EXCLUDE_EMAILS = [
    'councillor_saxe@toronto.ca',
    'dianne.saxe@toronto.ca',
    'dsaxe@envirolaw.com',
    'bob@diannesaxe.ca',
    'sydney@diannesaxe.ca',
    'christian.cullis@toronto.ca',
    'noreply@',
    'no-reply@',
    'donotreply@',
    'mailer-daemon@',
    'postmaster@',
]

EXCLUDE_DOMAINS = [
    '@toronto.ca',  
    '@diannesaxe.ca',  
    '@salesforce.com',
    '@microsoft.com',
    '@envirolaw.com',
]

# Domains that are likely automated/bulk and can be flagged
BULK_SENDER_DOMAINS = [
    '@actionnetwork.org',
    '@change.org',
    '@campaigns.',
    '@petition',
    '@mailchimp',
    '@constantcontact',
]



# Email normalization

def normalize_email(email: str) -> Optional[str]:
    """Normalize email for comparison."""
    if not email or pd.isna(email):
        return None
    email = str(email).lower().strip()
    if not re.match(r'^[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}$', email):
        return None
    return email


def is_excluded_email(email: str) -> bool:
    """Check if email should be excluded from analysis."""
    if not email:
        return True
    email = email.lower()
    
    # Check exact matches
    for excl in EXCLUDE_EMAILS:
        if excl in email:
            return True
    
    # Check domain exclusions
    for domain in EXCLUDE_DOMAINS:
        if domain in email:
            return True
    
    return False


def is_bulk_sender(email: str) -> bool:
    """Check if email is likely from a bulk/petition sender."""
    if not email:
        return False
    email = email.lower()
    return any(domain in email for domain in BULK_SENDER_DOMAINS)


#data loading

def load_salesforce_emails(filepath: str) -> Tuple[Set[str], pd.DataFrame]:
    """
    Load Salesforce export and extract all known constituent emails.
    Returns (set of emails, dataframe for reference).
    """
    print(f"Loading Salesforce data from {filepath}...")
    df = pd.read_csv(filepath, encoding='utf-8-sig')
    
    # Collect all emails from various columns
    email_columns = ['Contact: Email', 'Web Email', 'Extracted_Email']
    emails = set()
    
    for col in email_columns:
        if col in df.columns:
            for email in df[col].dropna():
                normalized = normalize_email(email)
                if normalized and not is_excluded_email(normalized):
                    emails.add(normalized)
    
    print(f"Found {len(emails)} unique constituent emails in Salesforce")
    return emails, df


def detect_outlook_columns(df: pd.DataFrame) -> dict:
    """Auto-detect column names in Outlook export."""
    column_map = {
        'email': None,
        'subject': None,
        'date': None,
        'body': None,
        'name': None,
    }
    
    # Common column name patterns
    patterns = {
        'email': ['senderemail', 'sender email', 'from', 'email', 'fromemail', 'from_email', 'senderemailaddress'],
        'subject': ['subject', 'subj', 'title', 'emailsubject'],
        'date': ['receiveddate', 'received', 'date', 'datetime', 'sentdate', 'sent', 'receivedtime'],
        'body': ['body', 'content', 'message', 'emailbody', 'text'],
        'name': ['sendername', 'sender name', 'from_name', 'fromname', 'name'],
    }
    
    for col in df.columns:
        col_lower = col.lower().replace(' ', '').replace('_', '')
        for key, candidates in patterns.items():
            if column_map[key] is None:
                for candidate in candidates:
                    if candidate.replace(' ', '').replace('_', '') in col_lower:
                        column_map[key] = col
                        break
    
    return column_map


def load_outlook_export(filepath: str) -> pd.DataFrame:
    """Load Outlook export (CSV or Excel from Access)."""
    print(f"Loading Outlook export from {filepath}...")
    
    # Detect file type
    if filepath.endswith('.xlsx') or filepath.endswith('.xls'):
        df = pd.read_excel(filepath)
    else:
        # Try different encodings
        for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                df = pd.read_csv(filepath, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
    
    print(f"Loaded {len(df)} emails from Outlook")
    print(f"Columns found: {list(df.columns)}")
    
    # Detect columns
    col_map = detect_outlook_columns(df)
    print(f"Detected columns: {col_map}")
    
    # Standardize column names
    if col_map['email']:
        df['_sender_email'] = df[col_map['email']].apply(normalize_email)
    else:
        raise ValueError("Could not detect sender email column in Outlook export")
    
    if col_map['subject']:
        df['_subject'] = df[col_map['subject']]
    else:
        df['_subject'] = ''
    
    if col_map['date']:
        df['_date'] = pd.to_datetime(df[col_map['date']], errors='coerce')
    else:
        df['_date'] = pd.NaT
    
    if col_map['name']:
        df['_sender_name'] = df[col_map['name']]
    else:
        df['_sender_name'] = ''
    
    return df


# Matching logic

def match_outlook_to_salesforce(
    outlook_df: pd.DataFrame,
    salesforce_emails: Set[str],
    salesforce_df: pd.DataFrame
) -> pd.DataFrame:
    """
    Compare Outlook emails against Salesforce records.
    Returns DataFrame with match status for each email.
    """
    results = []
    
    for idx, row in outlook_df.iterrows():
        email = row.get('_sender_email')
        subject = row.get('_subject', '')
        date = row.get('_date')
        name = row.get('_sender_name', '')
        
        result = {
            'sender_email': email,
            'sender_name': name,
            'subject': subject,
            'date': date,
            'status': 'unknown',
            'notes': '',
        }
        
        # Skip excluded emails
        if not email or is_excluded_email(email):
            result['status'] = 'excluded'
            result['notes'] = 'Staff/system email'
            results.append(result)
            continue
        
        # Check if email exists in Salesforce
        if email in salesforce_emails:
            result['status'] = 'in_salesforce'
            result['notes'] = 'Email found in Salesforce contacts'
        else:
            result['status'] = 'not_in_salesforce'
            result['notes'] = 'Email NOT found in any Salesforce record'
            
            # Flag bulk senders
            if is_bulk_sender(email):
                result['notes'] += ' | Likely bulk/petition sender'
        
        results.append(result)
    
    return pd.DataFrame(results)


# output

def create_output(
    results_df: pd.DataFrame,
    outlook_df: pd.DataFrame,
    output_path: str
):
    """Create Excel output with multiple sheets."""
    
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    status_counts = results_df['status'].value_counts()
    ws_summary['A1'] = "Status"
    ws_summary['B1'] = "Count"
    ws_summary['A1'].font = Font(bold=True)
    ws_summary['B1'].font = Font(bold=True)
    
    for i, (status, count) in enumerate(status_counts.items(), 2):
        ws_summary[f'A{i}'] = status
        ws_summary[f'B{i}'] = count
    
    # Colors for status
    colors = {
        'in_salesforce': '90EE90',      # Light green
        'not_in_salesforce': 'FFB6C1',  # Light red
        'excluded': 'D3D3D3',           # Light gray
    }
    
    # Not in Salesforce sheet (the main output)
    missed = results_df[results_df['status'] == 'not_in_salesforce'].copy()
    
    # Deduplicate by email (keep first occurrence)
    missed_deduped = missed.drop_duplicates(subset=['sender_email'], keep='first')
    
    ws_missed = wb.create_sheet("Not In Salesforce")
    ws_missed['A1'] = f"Unique emails not found in Salesforce: {len(missed_deduped)}"
    ws_missed['A1'].font = Font(bold=True, size=14)
    
    # Write headers
    headers = ['Sender Email', 'Sender Name', 'Subject', 'Date', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws_missed.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='DDDDDD')
    
    # Write data
    for r_idx, (_, row) in enumerate(missed_deduped.iterrows(), 4):
        ws_missed.cell(row=r_idx, column=1, value=row['sender_email'])
        ws_missed.cell(row=r_idx, column=2, value=row['sender_name'])
        ws_missed.cell(row=r_idx, column=3, value=str(row['subject'])[:100])
        ws_missed.cell(row=r_idx, column=4, value=str(row['date']))
        ws_missed.cell(row=r_idx, column=5, value=row['notes'])
    
    # Adjust column widths
    ws_missed.column_dimensions['A'].width = 35
    ws_missed.column_dimensions['B'].width = 25
    ws_missed.column_dimensions['C'].width = 50
    ws_missed.column_dimensions['D'].width = 20
    ws_missed.column_dimensions['E'].width = 40
    
    # All results sheet
    ws_all = wb.create_sheet("All Emails")
    for r_idx, row in enumerate(dataframe_to_rows(results_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
            elif r_idx > 1:
                status = results_df.iloc[r_idx-2]['status'] if r_idx-2 < len(results_df) else None
                if status in colors:
                    cell.fill = PatternFill('solid', fgColor=colors[status])
    
    wb.save(output_path)
    print(f"\nOutput saved to {output_path}")
    print(f"  - {len(missed_deduped)} unique emails not found in Salesforce")
    print(f"  - {len(results_df[results_df['status'] == 'in_salesforce'])} emails already in Salesforce")
    print(f"  - {len(results_df[results_df['status'] == 'excluded'])} excluded (staff/system)")


# main

def main():
    parser = argparse.ArgumentParser(
        description='Match Outlook emails against Salesforce to find missed cases'
    )
    parser.add_argument('salesforce', help='Salesforce case export CSV')
    parser.add_argument('outlook', help='Outlook email export (CSV or Excel)')
    parser.add_argument('output', help='Output Excel file')
    parser.add_argument('--include-extracted', action='store_true',
                       help='Also use Extracted_Email column from constituent_extractor output')
    
    args = parser.parse_args()
    
    # Load data
    sf_emails, sf_df = load_salesforce_emails(args.salesforce)
    outlook_df = load_outlook_export(args.outlook)
    
    # Match
    results = match_outlook_to_salesforce(outlook_df, sf_emails, sf_df)
    
    # Output
    create_output(results, outlook_df, args.output)


if __name__ == '__main__':
    main()
