"""
Outlook to Salesforce Matcher v2
=================================
Compares exported Outlook emails against Salesforce contacts and cases
to identify constituents who emailed the office but were never entered
into the system.

Improvements over v1:
- CSV/XLSX auto-detect for all inputs
- Fuzzy email matching (catches typos and alias variations)
- Frequency analysis (surfaces repeat senders first)
- Cross-references v4 extractor output (optional, avoids duplicate work)
- Date range filtering
- Sender categorisation (constituent / bulk / organisation / unknown)
- Better dedup with first-seen / last-seen / message count

Usage:
    python outlook_matcher_v2.py salesforce_cases.xlsx contacts.csv outlook_export.csv -o output.xlsx
    python outlook_matcher_v2.py cases.xlsx contacts.csv outlook.csv -o output.xlsx --after 2025-01-01
    python outlook_matcher_v2.py cases.xlsx contacts.csv outlook.csv -o output.xlsx --extractor-output v4_results.xlsx
"""

import argparse
import os
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional, Set, Dict, Tuple, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    from rapidfuzz import fuzz
    HAS_FUZZ = True
except ImportError:
    HAS_FUZZ = False


# config

# Staff/system emails to always exclude
EXCLUDE_EMAILS = {
    'councillor_saxe@toronto.ca',
    'dianne.saxe@toronto.ca',
    'dsaxe@envirolaw.com',
    'bob@diannesaxe.ca',
    'sydney@diannesaxe.ca',
    'christian.cullis@toronto.ca',
    'anne@diannesaxe.ca',
    'william.hopkins2@toronto.ca',
    'saxeforcouncil@gmail.com',
}

EXCLUDE_DOMAINS = {
    '@toronto.ca',
    '@diannesaxe.ca',
    '@salesforce.com',
    '@microsoft.com',
    '@envirolaw.com',
}

# Domains/patterns for automated or bulk senders
BULK_PATTERNS = [
    '@actionnetwork.org', '@change.org', '@campaigns.',
    '@petition', '@mailchimp', '@constantcontact',
    'noreply@', 'no-reply@', 'donotreply@', 'do-not-reply@',
    'mailer-daemon@', 'postmaster@', 'notifications@',
    'updates@', 'info@', 'newsletter@', 'support@',
    'bounce@', 'autoresponder@',
]

# Domains that suggest an organisation, not an individual
ORG_DOMAIN_HINTS = [
    '.gc.ca', '.gov.on.ca', '.on.ca', '.edu',
    '.org', '.ca',  # generic .ca not .rogers.ca etc
]

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


# utilities

def load_file(path: str) -> pd.DataFrame:
    """Load CSV or XLSX automatically. Tries both formats as fallback."""
    p = Path(path)
    errors = []

    if p.suffix.lower() in ('.xlsx', '.xls'):
        try:
            return pd.read_excel(path)
        except Exception as e:
            errors.append(f"Excel: {e}")
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                return pd.read_csv(path, encoding=enc)
            except Exception:
                continue
    else:
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                return pd.read_csv(path, encoding=enc)
            except Exception as e:
                errors.append(f"CSV({enc}): {e}")
                continue
        try:
            return pd.read_excel(path)
        except Exception as e:
            errors.append(f"Excel fallback: {e}")

    raise ValueError(
        f"Could not read {path}\n"
        f"  Tried formats: {'; '.join(errors[:3])}\n"
        f"  Hint: Check the file exists and the extension matches the format."
    )


def normalize_email(email) -> Optional[str]:
    """Normalize email for comparison."""
    if not email or pd.isna(email):
        return None
    e = str(email).lower().strip()
    if not re.match(r'^[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}$', e):
        return None
    return e


def is_excluded(email: str) -> bool:
    """Check if email should be excluded (staff/system)."""
    if not email:
        return True
    el = email.lower()
    if el in EXCLUDE_EMAILS:
        return True
    return any(d in el for d in EXCLUDE_DOMAINS)


def is_bulk(email: str) -> bool:
    """Check if email is likely automated/bulk."""
    if not email:
        return False
    el = email.lower()
    return any(p in el for p in BULK_PATTERNS)


def classify_sender(email: str) -> str:
    """Classify an email address: constituent / bulk / org / unknown."""
    if not email:
        return 'unknown'
    el = email.lower()

    if is_bulk(el):
        return 'bulk'

    domain = el.split('@')[-1] if '@' in el else ''

    # Common personal email providers → likely constituent
    personal_domains = {
        'gmail.com', 'yahoo.com', 'yahoo.ca', 'hotmail.com', 'outlook.com',
        'live.com', 'icloud.com', 'me.com', 'aol.com', 'protonmail.com',
        'mail.com', 'rogers.com', 'bell.net', 'sympatico.ca', 'cogeco.ca',
        'shaw.ca', 'telus.net',
    }
    if domain in personal_domains:
        return 'constituent'

    # University/government → org
    if any(domain.endswith(h) for h in ['.gc.ca', '.gov.on.ca', '.edu', '.ac.uk']):
        return 'org'

    # Everything else → default to constituent (could be personal domain)
    return 'constituent'


# contact index

class ContactIndex:
    """Index of all known emails from Salesforce contacts + cases."""

    def __init__(self, contacts_df: Optional[pd.DataFrame], cases_df: pd.DataFrame):
        self.emails: Set[str] = set()
        self.email_to_name: Dict[str, str] = {}
        self.email_to_contact_id: Dict[str, str] = {}

        # From contacts export
        if contacts_df is not None:
            for _, row in contacts_df.iterrows():
                email = normalize_email(row.get('Email'))
                if email and not is_excluded(email):
                    self.emails.add(email)
                    cid = str(row.get('Contact ID', ''))
                    name_parts = [str(row.get('First Name', '')), str(row.get('Last Name', ''))]
                    name = ' '.join(p for p in name_parts if p and p != 'nan').strip()
                    self.email_to_name[email] = name
                    self.email_to_contact_id[email] = cid

        # From cases (linked contacts)
        for _, row in cases_df.iterrows():
            cid = str(row.get('Contact ID', ''))
            if not cid or cid == 'nan' or cid == '000000000000000':
                continue
            name = str(row.get('Contact Name', ''))
            if name == 'nan':
                name = ''

            for col in ['Contact: Email', 'Web Email']:
                email = normalize_email(row.get(col))
                if email and not is_excluded(email):
                    self.emails.add(email)
                    if email not in self.email_to_name or not self.email_to_name[email]:
                        self.email_to_name[email] = name
                    if email not in self.email_to_contact_id:
                        self.email_to_contact_id[email] = cid

    def is_known(self, email: str) -> bool:
        """Exact match."""
        return email in self.emails

    def fuzzy_match(self, email: str) -> Optional[Tuple[str, int]]:
        """Fuzzy match against known emails. Returns (matched_email, score) or None."""
        if not HAS_FUZZ or not email:
            return None

        domain = email.split('@')[-1]
        prefix = email.split('@')[0]
        best_score = 0
        best_match = None

        for known in self.emails:
            if known.split('@')[-1] != domain:
                continue
            score = fuzz.ratio(prefix, known.split('@')[0])
            if score > best_score and score >= 85:
                best_score = score
                best_match = known

        return (best_match, best_score) if best_match else None


# outlook column detection

def detect_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    """Auto-detect column names in Outlook export."""
    mapping = {'email': None, 'subject': None, 'date': None, 'body': None, 'name': None}

    patterns = {
        'email': ['senderemail', 'sendermailaddress', 'senderemailaddress',
                   'from', 'email', 'fromemail', 'from_email'],
        'subject': ['subject', 'subj', 'title', 'emailsubject'],
        'date': ['receiveddate', 'received', 'date', 'datetime',
                 'sentdate', 'sent', 'receivedtime'],
        'body': ['body', 'content', 'message', 'emailbody', 'text'],
        'name': ['sendername', 'sendname', 'from_name', 'fromname', 'name', 'sender'],
    }

    for col in df.columns:
        normalized = col.lower().replace(' ', '').replace('_', '')
        for key, candidates in patterns.items():
            if mapping[key] is None:
                for c in candidates:
                    if c.replace(' ', '').replace('_', '') in normalized:
                        mapping[key] = col
                        break

    return mapping


# main matching

def match_outlook(
    outlook_df: pd.DataFrame,
    index: ContactIndex,
    extractor_emails: Optional[Set[str]] = None,
    after: Optional[datetime] = None,
    before: Optional[datetime] = None,
) -> pd.DataFrame:
    """
    Match each Outlook email against Salesforce.

    Returns a per-sender summary DataFrame with:
    - status: in_salesforce / fuzzy_match / in_extractor / not_in_salesforce / excluded / bulk
    - message count, first seen, last seen
    """
    cols = detect_columns(outlook_df)

    if not cols['email']:
        raise ValueError(
            f"Could not detect sender email column. Columns found: {list(outlook_df.columns)}"
        )

    print(f"  Detected columns: {cols}")

    # Normalise columns
    outlook_df['_email'] = outlook_df[cols['email']].apply(normalize_email)
    outlook_df['_subject'] = outlook_df[cols['subject']].fillna('') if cols['subject'] else ''
    outlook_df['_name'] = outlook_df[cols['name']].fillna('') if cols['name'] else ''

    if cols['date']:
        outlook_df['_date'] = pd.to_datetime(outlook_df[cols['date']], errors='coerce')
    else:
        outlook_df['_date'] = pd.NaT

    # Date filtering
    if after:
        outlook_df = outlook_df[outlook_df['_date'].isna() | (outlook_df['_date'] >= after)]
    if before:
        outlook_df = outlook_df[outlook_df['_date'].isna() | (outlook_df['_date'] <= before)]

    # Group by sender
    sender_stats: Dict[str, dict] = {}

    for _, row in outlook_df.iterrows():
        email = row['_email']
        if not email:
            continue

        if email not in sender_stats:
            sender_stats[email] = {
                'email': email,
                'name': str(row['_name']) if str(row['_name']) != 'nan' else '',
                'count': 0,
                'first_seen': row['_date'],
                'last_seen': row['_date'],
                'subjects': [],
            }

        s = sender_stats[email]
        s['count'] += 1
        if pd.notna(row['_date']):
            if pd.isna(s['first_seen']) or row['_date'] < s['first_seen']:
                s['first_seen'] = row['_date']
            if pd.isna(s['last_seen']) or row['_date'] > s['last_seen']:
                s['last_seen'] = row['_date']
        if len(s['subjects']) < 3:
            subj = str(row['_subject'])[:80]
            if subj and subj != 'nan':
                s['subjects'].append(subj)
        # Keep the best name (longest non-empty)
        nm = str(row['_name']) if str(row['_name']) != 'nan' else ''
        if len(nm) > len(s['name']):
            s['name'] = nm

    # Classify each sender
    results = []
    for email, stats in sender_stats.items():
        rec = {
            'email': email,
            'sender_name': stats['name'],
            'message_count': stats['count'],
            'first_seen': stats['first_seen'],
            'last_seen': stats['last_seen'],
            'sample_subjects': ' | '.join(stats['subjects']),
            'status': '',
            'match_detail': '',
            'sender_type': classify_sender(email),
            'sf_contact_name': '',
            'sf_contact_id': '',
        }

        # Exclusion check
        if is_excluded(email):
            rec['status'] = 'excluded'
            rec['match_detail'] = 'Staff/system email'
            results.append(rec)
            continue

        if is_bulk(email):
            rec['status'] = 'bulk'
            rec['match_detail'] = 'Automated/bulk sender'
            results.append(rec)
            continue

        # Exact match
        if index.is_known(email):
            rec['status'] = 'in_salesforce'
            rec['match_detail'] = 'Exact email match'
            rec['sf_contact_name'] = index.email_to_name.get(email, '')
            rec['sf_contact_id'] = index.email_to_contact_id.get(email, '')
            results.append(rec)
            continue

        # Fuzzy match
        fuzzy = index.fuzzy_match(email)
        if fuzzy:
            matched_email, score = fuzzy
            rec['status'] = 'fuzzy_match'
            rec['match_detail'] = f'{score}% match → {matched_email}'
            rec['sf_contact_name'] = index.email_to_name.get(matched_email, '')
            rec['sf_contact_id'] = index.email_to_contact_id.get(matched_email, '')
            results.append(rec)
            continue

        # Check extractor output (already captured by v4 but not yet in Salesforce)
        if extractor_emails and email in extractor_emails:
            rec['status'] = 'in_extractor'
            rec['match_detail'] = 'Found in v4 extractor output (not yet in Salesforce)'
            results.append(rec)
            continue

        # Not found
        rec['status'] = 'not_in_salesforce'
        rec['match_detail'] = 'No match found'
        results.append(rec)

    df = pd.DataFrame(results)
    if len(df) > 0:
        df = df.sort_values(['status', 'message_count'], ascending=[True, False])
    return df


# excel output

STATUS_COLORS = {
    'in_salesforce': '90EE90',
    'fuzzy_match': 'FFFACD',
    'in_extractor': 'ADD8E6',
    'not_in_salesforce': 'FFB6C1',
    'excluded': 'D3D3D3',
    'bulk': 'E8E8E8',
}


def write_sheet(wb, name: str, df: pd.DataFrame, status_col: Optional[str] = None):
    """Write a formatted sheet."""
    ws = wb.create_sheet(name)
    rows = list(dataframe_to_rows(df, index=False, header=True))

    for r_idx, row_data in enumerate(rows, 1):
        for c_idx, value in enumerate(row_data, 1):
            # Convert timestamps to strings to avoid Excel issues
            if isinstance(value, pd.Timestamp):
                value = value.strftime('%Y-%m-%d %H:%M') if pd.notna(value) else ''
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == 1:
                cell.font = Font(bold=True, size=10, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='4472C4')
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Color by status
    if status_col and status_col in df.columns and len(rows) > 1:
        col_idx = list(df.columns).index(status_col) + 1
        for r_idx in range(2, len(rows) + 1):
            status = ws.cell(row=r_idx, column=col_idx).value
            if status in STATUS_COLORS:
                fill = PatternFill('solid', fgColor=STATUS_COLORS[status])
                for c_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = fill

    # Auto-width
    for c_idx in range(1, len(df.columns) + 1):
        letter = get_column_letter(c_idx)
        max_len = max(
            len(str(ws.cell(row=1, column=c_idx).value or '')),
            *(len(str(ws.cell(row=r, column=c_idx).value or '')[:60])
              for r in range(2, min(len(rows) + 1, 30)))
        ) if len(rows) > 1 else 10
        ws.column_dimensions[letter].width = min(max_len + 3, 45)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    return ws


def create_output(results_df: pd.DataFrame, output_path: str):
    """Create Excel output."""
    wb = Workbook()

    # --- Dashboard ---
    ws = wb.active
    ws.title = "Dashboard"
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15

    r = 1
    ws.cell(row=r, column=1, value="OUTLOOK MATCHER v2 — RESULTS").font = Font(bold=True, size=16)
    r += 2

    total = len(results_df)
    ws.cell(row=r, column=1, value="Total unique senders analysed")
    ws.cell(row=r, column=2, value=total)
    r += 2

    ws.cell(row=r, column=1, value="STATUS").font = Font(bold=True, size=14)
    ws.cell(row=r, column=2, value="Senders").font = Font(bold=True)
    ws.cell(row=r, column=3, value="Messages").font = Font(bold=True)
    ws.column_dimensions['C'].width = 15
    r += 1

    status_descriptions = {
        'not_in_salesforce': 'Not in Salesforce — needs review',
        'in_salesforce': 'Already in Salesforce',
        'fuzzy_match': 'Probable match (fuzzy) — verify',
        'in_extractor': 'Found by v4 extractor — pending import',
        'bulk': 'Bulk/automated sender',
        'excluded': 'Staff/system email',
    }

    for status in ['not_in_salesforce', 'fuzzy_match', 'in_extractor',
                    'in_salesforce', 'bulk', 'excluded']:
        sub = results_df[results_df['status'] == status]
        if len(sub) == 0:
            continue
        ws.cell(row=r, column=1, value=status_descriptions.get(status, status))
        ws.cell(row=r, column=2, value=len(sub))
        ws.cell(row=r, column=3, value=int(sub['message_count'].sum()))
        if status in STATUS_COLORS:
            ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=STATUS_COLORS[status])
        r += 1

    # --- Not In Salesforce sheet (the main actionable output) ---
    missed = results_df[results_df['status'] == 'not_in_salesforce'].copy()
    if len(missed) > 0:
        missed = missed.sort_values('message_count', ascending=False)
        display_cols = [
            'email', 'sender_name', 'sender_type', 'message_count',
            'first_seen', 'last_seen', 'sample_subjects',
        ]
        display_cols = [c for c in display_cols if c in missed.columns]
        write_sheet(wb, 'Not In Salesforce', missed[display_cols])

    # --- Fuzzy Matches (need human verification) ---
    fuzzy = results_df[results_df['status'] == 'fuzzy_match'].copy()
    if len(fuzzy) > 0:
        fuzzy = fuzzy.sort_values('message_count', ascending=False)
        display_cols = [
            'email', 'sender_name', 'match_detail',
            'sf_contact_name', 'sf_contact_id',
            'message_count', 'first_seen', 'last_seen',
        ]
        display_cols = [c for c in display_cols if c in fuzzy.columns]
        write_sheet(wb, 'Fuzzy Matches', fuzzy[display_cols])

    # --- In Extractor ---
    ext = results_df[results_df['status'] == 'in_extractor'].copy()
    if len(ext) > 0:
        ext = ext.sort_values('message_count', ascending=False)
        display_cols = ['email', 'sender_name', 'message_count', 'first_seen', 'last_seen']
        write_sheet(wb, 'In Extractor', ext[display_cols])

    # --- Frequent Senders (top 50 by message count, regardless of status) ---
    freq = results_df.nlargest(50, 'message_count').copy()
    display_cols = [
        'email', 'sender_name', 'status', 'message_count',
        'first_seen', 'last_seen', 'sample_subjects',
    ]
    display_cols = [c for c in display_cols if c in freq.columns]
    write_sheet(wb, 'Top Senders', freq[display_cols], status_col='status')

    # --- All Senders ---
    write_sheet(wb, 'All Senders', results_df, status_col='status')

    wb.save(output_path)


# main

def process(
    cases_path: str,
    contacts_path: Optional[str],
    outlook_path: str,
    output_path: str,
    extractor_path: Optional[str] = None,
    after: Optional[str] = None,
    before: Optional[str] = None,
):
    # Load Salesforce data
    print(f"Loading cases from {cases_path}...")
    cases_df = load_file(cases_path)
    print(f"  {len(cases_df)} cases")

    contacts_df = None
    if contacts_path:
        print(f"Loading contacts from {contacts_path}...")
        contacts_df = load_file(contacts_path)
        print(f"  {len(contacts_df)} contacts")

    # Build index
    print("Building contact index...")
    index = ContactIndex(contacts_df, cases_df)
    print(f"  {len(index.emails)} known emails indexed")

    # Load extractor output for cross-reference
    extractor_emails: Optional[Set[str]] = None
    if extractor_path:
        print(f"Loading extractor output from {extractor_path}...")
        ext_df = load_file(extractor_path)
        extractor_emails = set()
        for col in ['Best_Email', 'All_Emails']:
            if col in ext_df.columns:
                for val in ext_df[col].dropna():
                    for e in str(val).split(';'):
                        e = e.strip().lower()
                        if e and '@' in e and not is_excluded(e):
                            extractor_emails.add(e)
        print(f"  {len(extractor_emails)} emails from extractor")

    # Load Outlook
    print(f"Loading Outlook export from {outlook_path}...")
    outlook_df = load_file(outlook_path)
    print(f"  {len(outlook_df)} emails")

    # Parse date filters
    after_dt = pd.to_datetime(after) if after else None
    before_dt = pd.to_datetime(before) if before else None

    # Match
    print("Matching...")
    results = match_outlook(outlook_df, index, extractor_emails, after_dt, before_dt)

    # Stats
    print(f"\n{'='*60}")
    print("MATCHING COMPLETE")
    print(f"{'='*60}")
    print(f"\nTotal unique senders: {len(results)}")
    for status in ['not_in_salesforce', 'fuzzy_match', 'in_extractor',
                    'in_salesforce', 'bulk', 'excluded']:
        sub = results[results['status'] == status]
        if len(sub) > 0:
            msgs = int(sub['message_count'].sum())
            print(f"  {status:25s} {len(sub):>5} senders  ({msgs:>6} messages)")

    # Write
    print(f"\nWriting to {output_path}...")
    create_output(results, output_path)
    print(f"Output: {output_path}")

    return results


def main():
    parser = argparse.ArgumentParser(description='Outlook to Salesforce Matcher v2')
    parser.add_argument('cases', help='Salesforce cases export (CSV or XLSX)')
    parser.add_argument('contacts', nargs='?', help='Salesforce contacts export (CSV or XLSX)')
    parser.add_argument('outlook', help='Outlook email export (CSV or XLSX)')
    parser.add_argument('-o', '--output', default='outlook_match_results.xlsx', help='Output file')
    parser.add_argument('--extractor-output', help='v4 extractor output for cross-reference')
    parser.add_argument('--after', help='Only include emails after this date (YYYY-MM-DD)')
    parser.add_argument('--before', help='Only include emails before this date (YYYY-MM-DD)')

    args = parser.parse_args()
    process(
        cases_path=args.cases,
        contacts_path=args.contacts,
        outlook_path=args.outlook,
        output_path=args.output,
        extractor_path=args.extractor_output,
        after=args.after,
        before=args.before,
    )


if __name__ == '__main__':
    main()
