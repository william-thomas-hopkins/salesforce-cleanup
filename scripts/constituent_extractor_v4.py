"""
Constituent Contact Extractor v4
=================================
Clean rewrite with major improvements over v3:

1. EXTRACTION: Contact block detection (name/addr/phone grouped),
   structured [Parsed Address] parsing, signature block patterns,
   multi-person case detection with per-person grouping.

2. DISAMBIGUATION: Better multi-person handling, LLM used only when
   regex can't resolve. Primary constituent identification heuristics.

3. MATCHING: Accepts separate Contacts export for richer matching.
   Fuzzy email matching (Levenshtein). Phone-based matching.

4. OUTPUT: Priority-ordered sheets, multi-person cases separated,
   staff-oriented action descriptions, summary dashboard.

Usage:
    python constituent_extractor_v4.py cases.xlsx contacts.csv -o output.xlsx
    python constituent_extractor_v4.py cases.xlsx contacts.csv -o output.xlsx --llm
"""

import argparse
import json
import os
import re
import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Set

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz
    HAS_FUZZ = True
except ImportError:
    HAS_FUZZ = False

try:
    import anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# config

TEXT_FIELDS = ['Subject', 'Description', 'Case Notes', 'Case Comments']

STAFF_EMAILS = {
    'bob@diannesaxe.ca', 'sydney@diannesaxe.ca', 'christian.cullis@toronto.ca',
    'councillor_saxe@toronto.ca', 'dianne.saxe@toronto.ca', 'dsaxe@envirolaw.com',
    'saxeforcouncil@gmail.com', 'anne@diannesaxe.ca', 'william.hopkins2@toronto.ca',
}

STAFF_DOMAINS = ['@diannesaxe.ca', '@toronto.ca', '@envirolaw.com']

VALID_AREA_CODES = {
    '416', '647', '437', '905', '289', '365',
    '519', '226', '548', '613', '343', '705', '249', '807',
}

# regex patterns

RE_EMAIL = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', re.I)
RE_PHONE_10 = re.compile(r'\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4}')
RE_PHONE_7 = re.compile(r'(?<!\d)\d{3}[\s.\-]?\d{4}(?!\d)')
RE_POSTAL = re.compile(r'\b[Mm]\d[A-Za-z][\s]?\d[A-Za-z]\d\b')

STREET_SUFFIXES = (
    r'(?:St(?:reet)?|Ave(?:nue)?|Rd|Road|Dr(?:ive)?|Blvd|Boulevard|'
    r'Cres(?:cent)?|Ct|Court|Pl(?:ace)?|Terr(?:ace)?|Lane|Ln|Way|'
    r'Circle|Cir|Trail|Trl|Park|Pk|Gardens?|Gate|Grove|Heights?|'
    r'Hts|Hill|Square|Sq|Row|Walk|Mews)'
)

ADDRESS_PATTERNS = [
    # Standard: 123 Main St, 123A Main Street West
    re.compile(
        rf'(\d+[A-Za-z]?(?:\s*[-–]\s*\d+[A-Za-z]?)?)\s+'
        rf'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+'
        rf'{STREET_SUFFIXES}\.?'
        rf'(?:\s*(?:East|West|North|South|E\.?|W\.?|N\.?|S\.?))?',
        re.I
    ),
    # ALL CAPS: 690 MANNING AVE
    re.compile(
        rf'(\d+[A-Za-z]?)\s+([A-Z]{{2,}}(?:\s+[A-Z]{{2,}})*)\s+{STREET_SUFFIXES}',
        re.I
    ),
    # St. Name patterns: St. George St, St. Clair Ave
    re.compile(
        rf'(\d+[A-Za-z]?)\s+(St\.?\s+[A-Z][a-z]+)\s+(?:Street|St\.?|Ave(?:nue)?|Rd|Road)',
        re.I
    ),
    # lowercase: 120 macpherson ave
    re.compile(
        rf'(\d+[A-Za-z]?)\s+([a-z]+(?:\s+[a-z]+)*)\s+{STREET_SUFFIXES}',
        re.I
    ),
    # With unit prefix: Unit 5, 123 Main St
    re.compile(
        rf'(?:Unit|Apt|Suite|#)\s*\d+[A-Za-z]?\s*,?\s*(\d+)\s+'
        rf'([A-Za-z]+(?:\s+[A-Za-z]+)*)\s+{STREET_SUFFIXES}',
        re.I
    ),
    # Loose: just "400 Markham" without suffix (common in notes)
    re.compile(
        r'(\d+)\s+([A-Z][a-z]{2,}(?:\s+[A-Z][a-z]+)?)\s*(?:\.|,|$|\r?\n)',
        re.M
    ),
]

ADDRESS_FALSE_POSITIVES = {
    'minute', 'minutes', 'hour', 'hours', 'day', 'days', 'week', 'weeks',
    'year', 'years', 'month', 'months', 'meter', 'meters', 'dogs', 'cats',
    'cars', 'people', 'times', 'units', 'feet', 'inches', 'miles', 'block',
    'children', 'students', 'residents', 'members', 'guests', 'percent',
    'dollars', 'items', 'calls', 'emails', 'complaints',
}

# Structured address in Case Notes: [Parsed Address] street="...", city="...", postal="..."
RE_PARSED_ADDRESS = re.compile(
    r'\[Parsed Address\]\s*'
    r'street="([^"]*)"'
    r'(?:,\s*city="([^"]*)")?'
    r'(?:,\s*postal="([^"]*)")?'
)

# Email chain indicators
EMAIL_CHAIN_PATTERNS = [
    re.compile(r'^From:\s*', re.M | re.I),
    re.compile(r'^Sent:\s*', re.M | re.I),
    re.compile(r'[-]{3,}.*forwarded.*[-]{3,}', re.I),
    re.compile(r'[-]{3,}.*original message.*[-]{3,}', re.I),
    re.compile(r'^>+\s*', re.M),
    re.compile(r'wrote:$', re.M | re.I),
]

# Constituent address context phrases
CONSTITUENT_ADDR_PHRASES = [
    r'i\s+live\s+at', r'i\s+reside\s+at',
    r'i\s+am\s+(?:a\s+)?resident\s+(?:of|at)',
    r'my\s+(?:home\s+)?address\s+is', r'we\s+live\s+at',
    r'our\s+(?:home\s+)?address', r'my\s+property\s+at',
    r'i\s+own\s+(?:a\s+)?(?:home|house|property)\s+at',
]

# Title prefixes to strip from names
TITLE_PREFIX = r'(?:Mr\.?|Mrs\.?|Ms\.?|Dr\.?|Prof\.?)\s+'

# Name patterns
NAME_PATTERNS = [
    # "My name is [Mr.] John Smith"
    (re.compile(rf'my\s+name\s+is\s+(?:{TITLE_PREFIX})?([A-Z][a-z]+(?:\s+[A-Z][a-z\-\']+)+)', re.I), 'intro'),
    (re.compile(r"(?:^|[.!?]\s+)I'?m\s+([A-Z][a-z]+\s+[A-Z][a-z\-\']+)(?=[\s,.])", re.M), 'intro'),
    (re.compile(r'this\s+is\s+([A-Z][a-z]+\s+[A-Z][a-z\-\']+)(?:\s+(?:calling|writing|from))', re.I), 'intro'),
    (re.compile(
        r'(?:Yours\s+truly|Sincerely|Best\s+regards?|Thanks|Thank\s+you|Regards|Cheers|Best|Warmly)'
        r'[,.\s]*\r?\n+\s*([A-Z][a-z]+(?:\s+[A-Z][a-z\-\']+)+)'
        r'(?=\s*(?:\r?\n|$|[,\r]))', re.I
    ), 'signoff'),
    (re.compile(r'^\s*([A-Z][a-z]+\s+[A-Z][a-z\-\']+)\s*$\n.*?(?:@|\(\d{3}\)|\d{3}[\-.])', re.M), 'contact_block'),
    (re.compile(r'\r?\n\s*([A-Z][a-z]+(?:\s+[A-Z][a-z\-\']+)+)\s*$'), 'end'),
]

# Name line regex for person block detection — allows title prefixes
RE_NAME_LINE = re.compile(
    rf'^\s*(?:{TITLE_PREFIX})?([A-Z][a-z]+(?:\s+[A-Z][a-z\-\']+)+)\s*$'
)

FALSE_POSITIVE_NAMES = {
    'dear councillor', 'dear dianne', 'hello councillor', 'hi councillor',
    'dear ms', 'dear mr', 'dear dr', 'hello dianne', 'hi dianne',
    'councillor saxe', 'dianne saxe', 'dr saxe', 'ms saxe',
    'city of', 'university of', 'town of', 'province of',
    'best regards', 'kind regards', 'warm regards', 'on behalf',
    'sent from', 'get outlook', 'hi there', 'hello there',
    'good morning', 'good afternoon', 'good evening',
}


# data structures


@dataclass
class PersonBlock:
    """A grouped set of contact info for one person found in text."""
    name: Optional[str] = None
    emails: List[str] = field(default_factory=list)
    phones: List[Tuple[str, str]] = field(default_factory=list)  # (status, normalized)
    addresses: List[str] = field(default_factory=list)
    postals: List[str] = field(default_factory=list)
    is_constituent: bool = False
    source: str = ""  # 'contact_block', 'signature', 'body', 'parsed'
    line_range: Tuple[int, int] = (0, 0)


@dataclass
class CaseExtraction:
    """Complete extraction results for a single case."""
    case_number: str = ""
    case_id: str = ""

    # Salesforce existing data
    sf_contact_id: str = ""
    sf_contact_name: str = ""
    sf_contact_email: str = ""
    sf_web_email: str = ""
    sf_phone: str = ""
    sf_has_contact: bool = False
    sf_is_contaminated: bool = False

    # Parsed structured data (from [Parsed Address] in Case Notes)
    parsed_street: str = ""
    parsed_city: str = ""
    parsed_postal: str = ""

    # People found in text
    persons: List[PersonBlock] = field(default_factory=list)

    # All raw extractions (union across all persons + ungrouped)
    all_emails: List[str] = field(default_factory=list)
    all_phones: List[Tuple[str, str]] = field(default_factory=list)
    all_addresses: List[str] = field(default_factory=list)
    all_postals: List[Tuple[bool, str]] = field(default_factory=list)
    all_names: List[Tuple[str, str]] = field(default_factory=list)

    # Best guess for THE constituent
    best_email: str = ""
    best_phone: str = ""
    best_address: str = ""
    best_postal: str = ""
    best_name: str = ""

    # LLM results
    llm_email: str = ""
    llm_phone: str = ""
    llm_address: str = ""
    llm_complaint_addr: str = ""
    llm_confidence: str = ""
    llm_reasoning: str = ""
    llm_used: bool = False

    # Matching
    matched_contact_id: str = ""
    matched_contact_name: str = ""
    match_method: str = ""  # 'exact_email', 'fuzzy_email', 'phone', 'web_email'

    # Flags
    flags: List[str] = field(default_factory=list)
    is_email_chain: bool = False
    is_multi_person: bool = False
    person_count: int = 0
    _combined_text: str = ""  # internal: combined text for LLM checks

    # Action
    action: str = ""
    confidence: str = ""
    missing_fields: List[str] = field(default_factory=list)
    extractable_fields: List[str] = field(default_factory=list)


# utility functions

def safe_str(value) -> str:
    if pd.isna(value) or value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() == 'nan' else s


def is_staff_email(email: str) -> bool:
    if not email:
        return False
    e = email.lower().strip()
    if e in STAFF_EMAILS:
        return True
    return any(d in e for d in STAFF_DOMAINS)


def load_file(path: str) -> pd.DataFrame:
    """Load CSV or XLSX automatically. Tries both formats as fallback."""
    p = Path(path)
    errors = []

    # Try XLSX first if extension suggests it, otherwise try CSV first
    if p.suffix.lower() in ('.xlsx', '.xls'):
        try:
            return pd.read_excel(path)
        except Exception as e:
            errors.append(f"Excel: {e}")
        # Fallback: try CSV
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                return pd.read_csv(path, encoding=enc)
            except Exception:
                continue
    else:
        # Try CSV first
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                return pd.read_csv(path, encoding=enc)
            except Exception as e:
                errors.append(f"CSV({enc}): {e}")
                continue
        # Fallback: try XLSX (file might be xlsx despite extension)
        try:
            return pd.read_excel(path)
        except Exception as e:
            errors.append(f"Excel fallback: {e}")

    raise ValueError(
        f"Could not read {path}\n"
        f"  Tried formats: {'; '.join(errors[:3])}\n"
        f"  Hint: Check the file exists and the extension matches the format."
    )


def normalize_phone(phone: str) -> Tuple[str, str]:
    """Returns (status, normalized). Status: valid|partial|invalid."""
    digits = re.sub(r'\D', '', phone)
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    if len(digits) == 10:
        if digits[:3] in VALID_AREA_CODES:
            return 'valid', f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        return 'valid', f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    if len(digits) == 7:
        return 'partial', f"{digits[:3]}-{digits[3:]}"
    return 'invalid', phone


def normalize_postal(postal: str) -> Tuple[bool, str]:
    """Returns (is_valid_toronto, normalized)."""
    n = postal.upper().replace(" ", "")
    if len(n) != 6 or not n.startswith('M'):
        return False, n
    if re.match(r'^M\d[A-Z]\d[A-Z]\d$', n):
        return True, f"{n[:3]} {n[3:]}"
    return False, n


def detect_email_chain(text: str) -> bool:
    if not text:
        return False
    return any(p.search(text) for p in EMAIL_CHAIN_PATTERNS)


def is_constituent_context(text: str, addr: str) -> bool:
    """Check if address appears near 'I live at' type phrases."""
    if not text or not addr:
        return False
    tl = text.lower()
    al = addr.lower()[:20]
    pos = tl.find(al)
    if pos == -1:
        return False
    ctx = tl[max(0, pos - 150):pos]
    return any(re.search(p, ctx) for p in CONSTITUENT_ADDR_PHRASES)


def clean_name(name: str) -> Optional[str]:
    """Validate and clean an extracted name."""
    if not name:
        return None
    name = re.sub(r'\s*[\r\n].*$', '', name)
    name = re.sub(r'\s*\d+.*$', '', name)
    for pat in [r'\s+on behalf.*$', r'\s+sent from.*$', r'\s+get outlook.*$', r'\s+\(.*\)$']:
        name = re.sub(pat, '', name, flags=re.I)
    name = name.strip(' ,.')
    if name.lower() in FALSE_POSITIVE_NAMES:
        return None
    words = name.split()
    if len(words) < 2 or len(words) > 5:
        return None
    if not all(len(w) >= 2 for w in words):
        return None
    if not all(w[0].isupper() for w in words):
        return None
    return name


# extraction engine

def extract_emails(text: str) -> List[str]:
    """Extract non-staff emails from text."""
    if not text:
        return []
    return list(set(e for e in RE_EMAIL.findall(text) if not is_staff_email(e)))


def extract_phones(text: str) -> List[Tuple[str, str]]:
    """Extract phones. Returns list of (status, normalized)."""
    if not text:
        return []
    phones = []
    seen = set()
    for m in RE_PHONE_10.finditer(text):
        s, n = normalize_phone(m.group())
        if s != 'invalid' and n not in seen:
            phones.append((s, n))
            seen.add(n)
    for m in RE_PHONE_7.finditer(text):
        s, n = normalize_phone(m.group())
        if s == 'partial' and n not in seen:
            phones.append((s, n))
            seen.add(n)
    return phones


def extract_addresses(text: str) -> List[str]:
    """Extract street addresses."""
    if not text:
        return []
    addrs = set()
    for pat in ADDRESS_PATTERNS:
        for m in pat.finditer(text):
            a = re.sub(r'\s+', ' ', m.group(0).strip())
            addrs.add(a)

    filtered = []
    for a in addrs:
        al = a.lower()
        if any(w in al for w in ADDRESS_FALSE_POSITIVES):
            continue
        if len(a) < 6:
            continue
        filtered.append(a)
    return list(set(filtered))


def extract_names(text: str) -> List[Tuple[str, str]]:
    """Extract names. Returns list of (name, pattern_type)."""
    if not text:
        return []
    names = []
    seen = set()
    for pat, ptype in NAME_PATTERNS:
        for m in pat.finditer(text):
            n = clean_name(m.group(1))
            if n and n.lower() not in seen:
                names.append((n, ptype))
                seen.add(n.lower())

    # Also check for "Mr./Mrs./Dr. Firstname Lastname" at start of lines
    title_pat = re.compile(
        rf'^(?:Mr\.?|Mrs\.?|Ms\.?|Dr\.?)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z\-\']+)+)',
        re.M
    )
    for m in title_pat.finditer(text):
        n = clean_name(m.group(1))
        if n and n.lower() not in seen:
            names.append((n, 'titled'))
            seen.add(n.lower())

    return names


def extract_postals(text: str) -> List[Tuple[bool, str]]:
    if not text:
        return []
    postals = []
    seen = set()
    for m in RE_POSTAL.finditer(text):
        valid, norm = normalize_postal(m.group())
        if norm and norm not in seen:
            postals.append((valid, norm))
            seen.add(norm)
    return postals


def parse_structured_address(case_notes: str) -> Tuple[str, str, str]:
    """Extract from [Parsed Address] structured data in Case Notes."""
    if not case_notes:
        return "", "", ""
    m = RE_PARSED_ADDRESS.search(case_notes)
    if not m:
        return "", "", ""
    street = m.group(1).strip() if m.group(1) else ""
    city = m.group(2).strip() if m.group(2) else ""
    postal = m.group(3).strip() if m.group(3) else ""
    return street, city, postal


# person block detection

def detect_person_blocks(text: str) -> List[PersonBlock]:
    """
    Detect grouped contact info belonging to individual people.

    Patterns detected:
    1. Contact block at start: Name\\nAddress\\nPhone
    2. Signature block at end: Regards,\\nName\\nPhone\\nEmail
    3. Call log entries: Name Time\\nAddress\\nPhone\\nNotes
    4. Labeled fields: Name: X, Email: Y, Phone: Z
    """
    if not text:
        return []

    lines = text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
    persons = []

    # --- Pattern 1: Contact block at start (first 5 non-empty lines) ---
    first_lines = []
    for i, line in enumerate(lines[:10]):
        if line.strip():
            first_lines.append((i, line.strip()))
        if len(first_lines) >= 5:
            break

    if len(first_lines) >= 2:
        name_match = RE_NAME_LINE.match(first_lines[0][1])
        if name_match:
            block = PersonBlock(source='contact_block_start')
            block.name = clean_name(name_match.group(1))
            block.line_range = (first_lines[0][0], first_lines[-1][0])
            block.is_constituent = True  # first person listed is usually the constituent

            for _, line in first_lines[1:]:
                for e in extract_emails(line):
                    block.emails.append(e)
                for p in extract_phones(line):
                    block.phones.append(p)
                addrs = extract_addresses(line)
                if addrs:
                    block.addresses.extend(addrs)
                elif re.match(r'^\d+\s+\w', line) and not any(w in line.lower() for w in ADDRESS_FALSE_POSITIVES):
                    block.addresses.append(line)
                postals = extract_postals(line)
                block.postals.extend(p for _, p in postals)

            if block.name and (block.phones or block.emails or block.addresses):
                persons.append(block)

    # --- Pattern 2: Signature block at end ---
    last_lines = []
    for i in range(len(lines) - 1, max(len(lines) - 15, -1), -1):
        if lines[i].strip():
            last_lines.insert(0, (i, lines[i].strip()))
        if len(last_lines) >= 8:
            break

    if len(last_lines) >= 2:
        # Look for name line followed or preceded by contact info
        for j, (li, line) in enumerate(last_lines):
            nm = RE_NAME_LINE.match(line)
            if not nm:
                continue
            cleaned = clean_name(nm.group(1))
            if not cleaned:
                continue

            # Check if nearby lines have contact info
            nearby = last_lines[max(0, j-1):j+4]
            block = PersonBlock(source='signature', name=cleaned)
            block.line_range = (last_lines[0][0], last_lines[-1][0])
            block.is_constituent = True

            for _, nl in nearby:
                if nl == line:
                    continue
                # Labeled: "Email: x@y.com", "Cell: 416-..."
                label_email = re.search(r'(?:email|e-mail)\s*:\s*(\S+@\S+)', nl, re.I)
                if label_email:
                    e = label_email.group(1).rstrip('.,;')
                    if not is_staff_email(e):
                        block.emails.append(e)
                    continue

                label_phone = re.search(r'(?:cell|phone|tel|mobile|fax)\s*:\s*([\d\(\)\s\.\-]+)', nl, re.I)
                if label_phone:
                    s, n = normalize_phone(label_phone.group(1))
                    if s != 'invalid':
                        block.phones.append((s, n))
                    continue

                for e in extract_emails(nl):
                    block.emails.append(e)
                for p in extract_phones(nl):
                    block.phones.append(p)
                addrs = extract_addresses(nl)
                block.addresses.extend(addrs)

            if block.phones or block.emails:
                # Don't duplicate if same person already found
                existing_names = {p.name.lower() for p in persons if p.name}
                if cleaned.lower() not in existing_names:
                    persons.append(block)
            break  # Only use first name match from end

    # --- Pattern 3: Call log / multi-person entries ---
    # Detect: "Name [optional Time]\nAddress\nPhone\nnotes"
    # Also: "Name\nPhone" or "Name\nAddress" blocks in middle of text
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue

        # Pattern A: "Name HH:MM" (call logs)
        call_match = re.match(
            r'^([A-Z][a-z]+(?:\s+[A-Z][a-z\-\']+)+)\s+\d{1,2}[;:]\d{2}\s*(?:am|pm|AM|PM)?\s*$',
            line
        )
        # Pattern B: Name on its own line in the middle (not first/last 3 lines, handled above)
        standalone_name = None
        if not call_match and i >= 1 and i < len(lines) - 3:
            standalone_name = RE_NAME_LINE.match(line)

        name_str = call_match.group(1) if call_match else (standalone_name.group(1) if standalone_name else None)

        if name_str:
            cname = clean_name(name_str)
            if cname:
                block = PersonBlock(source='call_log' if call_match else 'inline', name=cname)
                block.line_range = (i, min(i + 5, len(lines) - 1))
                has_contact = False

                for j in range(i + 1, min(i + 6, len(lines))):
                    nl = lines[j].strip()
                    if not nl:
                        break
                    if RE_NAME_LINE.match(nl) and j > i + 1:
                        break
                    for p in extract_phones(nl):
                        block.phones.append(p)
                        has_contact = True
                    addrs = extract_addresses(nl)
                    block.addresses.extend(addrs)
                    if addrs:
                        has_contact = True
                    elif re.match(r'^\d+\s+', nl) and not any(w in nl.lower() for w in ADDRESS_FALSE_POSITIVES):
                        block.addresses.append(nl)
                        has_contact = True
                    for e in extract_emails(nl):
                        block.emails.append(e)
                        has_contact = True

                if has_contact:
                    existing_names = {p.name.lower() for p in persons if p.name}
                    if cname.lower() not in existing_names:
                        persons.append(block)
        i += 1

    return persons


# matching engine

class ContactIndex:
    """Index of existing contacts for matching."""

    def __init__(self, contacts_df: Optional[pd.DataFrame], cases_df: pd.DataFrame):
        self.email_to_contact: Dict[str, Tuple[str, str]] = {}  # email -> (contact_id, contact_name)
        self.phone_to_contact: Dict[str, Tuple[str, str]] = {}  # phone_digits -> (contact_id, contact_name)
        self.all_contact_emails: Set[str] = set()

        # Build from contacts export
        if contacts_df is not None:
            for _, row in contacts_df.iterrows():
                cid = safe_str(row.get('Contact ID'))
                if not cid:
                    continue
                name_parts = [safe_str(row.get('First Name')), safe_str(row.get('Last Name'))]
                cname = ' '.join(p for p in name_parts if p)

                email = safe_str(row.get('Email')).lower()
                if email and not is_staff_email(email):
                    self.email_to_contact[email] = (cid, cname)
                    self.all_contact_emails.add(email)

                for col in ['Phone', 'Mobile']:
                    phone = safe_str(row.get(col))
                    if phone:
                        digits = re.sub(r'\D', '', phone)
                        if len(digits) >= 10:
                            digits = digits[-10:]
                            self.phone_to_contact[digits] = (cid, cname)

        # Also build from cases (linked cases may reference contacts not in contacts export)
        for _, row in cases_df.iterrows():
            cid = safe_str(row.get('Contact ID'))
            if not cid or cid == '000000000000000':
                continue
            cname = safe_str(row.get('Contact Name'))

            for col in ['Contact: Email', 'Web Email']:
                email = safe_str(row.get(col)).lower()
                if email and not is_staff_email(email):
                    if email not in self.email_to_contact:
                        self.email_to_contact[email] = (cid, cname)
                    self.all_contact_emails.add(email)

            phone = safe_str(row.get('Contact: Phone'))
            if phone:
                digits = re.sub(r'\D', '', phone)
                if len(digits) >= 10:
                    digits = digits[-10:]
                    if digits not in self.phone_to_contact:
                        self.phone_to_contact[digits] = (cid, cname)

    def match_email(self, email: str) -> Optional[Tuple[str, str, str]]:
        """Returns (contact_id, contact_name, method) or None."""
        if not email:
            return None
        el = email.lower().strip()

        # Exact match
        if el in self.email_to_contact:
            cid, cn = self.email_to_contact[el]
            return cid, cn, 'exact_email'

        # Fuzzy match
        if HAS_FUZZ:
            best_score = 0
            best_match = None
            for known in self.all_contact_emails:
                # Only compare within same domain
                if el.split('@')[-1] != known.split('@')[-1]:
                    continue
                score = fuzz.ratio(el.split('@')[0], known.split('@')[0])
                if score > best_score and score >= 85:
                    best_score = score
                    best_match = known

            if best_match:
                cid, cn = self.email_to_contact[best_match]
                return cid, cn, f'fuzzy_email({best_score}%:{best_match})'

        return None

    def match_phone(self, phone: str) -> Optional[Tuple[str, str, str]]:
        """Returns (contact_id, contact_name, method) or None."""
        if not phone:
            return None
        digits = re.sub(r'\D', '', phone)
        if len(digits) >= 10:
            digits = digits[-10:]
            if digits in self.phone_to_contact:
                cid, cn = self.phone_to_contact[digits]
                return cid, cn, 'phone'
        return None


# llm disambiguation

LLM_PROMPT = """You are extracting constituent contact info from a Toronto city councillor's case.

CASE TEXT:
{text}

EXTRACTED CANDIDATES:
- Emails: {emails}
- Phones: {phones}
- Addresses: {addresses}
- People detected: {people}

Identify which info belongs to THE CONSTITUENT (the person asking for help).
DO NOT select staff (@toronto.ca, @diannesaxe.ca), people being complained about,
or forwarded email senders (unless they ARE the constituent).

"I live at 400 Markham" = constituent address
"The property at 400 Markham is causing problems" = complaint address

Return ONLY valid JSON:
{{
  "constituent_email": "email or null",
  "constituent_phone": "phone or null",
  "constituent_address": "address or null",
  "complaint_address": "address or null",
  "confidence": "high/medium/low",
  "reasoning": "one sentence"
}}"""


def call_llm(text, emails, phones, addresses, people, client):
    try:
        prompt = LLM_PROMPT.format(
            text=text[:4000],
            emails=', '.join(emails) or 'None',
            phones=', '.join(f"{s}:{p}" for s, p in phones) or 'None',
            addresses=', '.join(addresses) or 'None',
            people=', '.join(
                f"{p.name}({p.source})" for p in people if p.name
            ) or 'None',
        )
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        txt = resp.content[0].text.strip()
        jm = re.search(r'\{[^{}]*\}', txt, re.DOTALL)
        if jm:
            return json.loads(jm.group())
    except Exception as e:
        return {'error': str(e)}
    return None


def needs_llm(ext: CaseExtraction) -> bool:
    """Determine if LLM disambiguation would help this case."""
    # Multiple non-staff emails — which one is the constituent's?
    if len(ext.all_emails) > 1:
        return True
    # Multi-person case — is our best guess the right person?
    if ext.is_multi_person:
        return True
    # Email chain with extracted emails but no linked contact
    if ext.is_email_chain and ext.all_emails and not ext.sf_has_contact:
        return True
    # Multiple addresses and no constituent context clue
    if len(ext.all_addresses) > 1:
        text_parts = []
        for a in ext.all_addresses:
            if is_constituent_context(ext._combined_text, a):
                return False  # We have a clear signal, no LLM needed
        return True
    return False


def _llm_priority_score(ext: CaseExtraction) -> int:
    """
    Score how valuable an LLM call would be for this case.
    Higher = more valuable = process first.

    Priority order:
    1. Multi-person unlinked cases (hardest to resolve by regex)
    2. Multiple emails, unlinked (LLM picks the right constituent email → CREATE_CONTACT)
    3. Email chains, unlinked (LLM disambiguates forwarded vs sender)
    4. Multiple addresses (LLM separates constituent addr from complaint addr)
    5. Already-linked cases get lower priority (enrichment, not linkage)
    """
    score = 0

    # Unlinked cases are more valuable than linked ones
    if not ext.sf_has_contact:
        score += 100

    # Multi-person is hardest to resolve
    if ext.is_multi_person:
        score += 50 + (ext.person_count * 10)

    # Multiple emails — LLM picks the right one
    if len(ext.all_emails) > 1:
        score += 40 + (len(ext.all_emails) * 5)

    # Email chain ambiguity
    if ext.is_email_chain:
        score += 30

    # Multiple addresses — constituent vs complaint
    if len(ext.all_addresses) > 1:
        score += 20 + (len(ext.all_addresses) * 3)

    # Cases with more text give the LLM more to work with
    text_len = len(ext._combined_text)
    if text_len > 500:
        score += 10
    if text_len > 1500:
        score += 5

    return score


# main extraction pipeline

def extract_case(row: pd.Series, contact_index: ContactIndex,
                 client=None, use_llm=False) -> CaseExtraction:
    ext = CaseExtraction()

    # --- Salesforce fields ---
    ext.case_number = safe_str(row.get('Case Number'))
    ext.case_id = safe_str(row.get('Case ID'))
    ext.sf_contact_id = safe_str(row.get('Contact ID'))
    ext.sf_contact_name = safe_str(row.get('Contact Name'))
    ext.sf_contact_email = safe_str(row.get('Contact: Email'))
    ext.sf_web_email = safe_str(row.get('Web Email'))
    ext.sf_phone = safe_str(row.get('Contact: Phone'))

    ext.sf_has_contact = bool(ext.sf_contact_id and ext.sf_contact_id != '000000000000000')
    # Only contaminated if the Contact: Email itself is staff.
    # Web Email being staff is normal (staff forwarded it) — not contamination if Contact Email is legit.
    ext.sf_is_contaminated = bool(
        ext.sf_contact_email and is_staff_email(ext.sf_contact_email)
        and not (ext.sf_web_email and not is_staff_email(ext.sf_web_email))  # has legit web email = not contaminated
    )

    # --- Combine text ---
    text_parts = []
    for f in TEXT_FIELDS:
        v = row.get(f)
        if pd.notna(v) and str(v).strip():
            text_parts.append(str(v))
    combined = "\n".join(text_parts)
    ext._combined_text = combined

    if not combined.strip():
        ext.action = "NO_TEXT"
        ext.confidence = "none"
        return ext

    # --- Structured [Parsed Address] ---
    case_notes = safe_str(row.get('Case Notes'))
    ext.parsed_street, ext.parsed_city, ext.parsed_postal = parse_structured_address(case_notes)
    if ext.parsed_street:
        ext.flags.append("HAS_PARSED_ADDR")

    # --- Email chain detection ---
    ext.is_email_chain = detect_email_chain(combined)
    if ext.is_email_chain:
        ext.flags.append("EMAIL_CHAIN")

    # --- Raw extractions ---
    ext.all_emails = extract_emails(combined)
    ext.all_phones = extract_phones(combined)
    ext.all_addresses = extract_addresses(combined)
    ext.all_postals = extract_postals(combined)
    ext.all_names = extract_names(combined)

    if len(ext.all_emails) > 1:
        ext.flags.append(f"MULTI_EMAIL({len(ext.all_emails)})")
    if len(ext.all_addresses) > 1:
        ext.flags.append(f"MULTI_ADDR({len(ext.all_addresses)})")

    # --- Person block detection ---
    ext.persons = detect_person_blocks(combined)
    ext.person_count = len(ext.persons)
    if ext.person_count > 1:
        ext.is_multi_person = True
        ext.flags.append(f"MULTI_PERSON({ext.person_count})")
    elif ext.person_count == 1:
        ext.flags.append("SINGLE_PERSON_BLOCK")

    # --- Determine best constituent contact info ---
    _resolve_best_contact(ext, combined)

    # --- LLM if needed ---
    if use_llm and client and needs_llm(ext):
        result = call_llm(combined, ext.all_emails, ext.all_phones,
                          ext.all_addresses, ext.persons, client)
        if result and 'error' not in result:
            ext.llm_used = True
            ext.llm_email = result.get('constituent_email') or ''
            ext.llm_phone = result.get('constituent_phone') or ''
            ext.llm_address = result.get('constituent_address') or ''
            ext.llm_complaint_addr = result.get('complaint_address') or ''
            ext.llm_confidence = result.get('confidence') or ''
            ext.llm_reasoning = result.get('reasoning') or ''
            ext.flags.append("LLM_USED")

            # LLM overrides best guess
            if ext.llm_email:
                ext.best_email = ext.llm_email
            if ext.llm_phone:
                ext.best_phone = ext.llm_phone
            if ext.llm_address:
                ext.best_address = ext.llm_address

    # --- Match to existing contacts ---
    if not ext.sf_has_contact:
        _match_to_contacts(ext, contact_index)

    # --- Determine missing/extractable ---
    _compute_fields(ext)

    # --- Determine action ---
    ext.action, ext.confidence = _determine_action(ext)

    return ext


def _resolve_best_contact(ext: CaseExtraction, text: str):
    """Pick the most likely constituent email/phone/address/name."""

    # --- EMAIL ---
    # Priority: Web Email > single extracted > person block constituent > first extracted
    web = ext.sf_web_email
    if web and not is_staff_email(web):
        ext.best_email = web
    elif len(ext.all_emails) == 1:
        ext.best_email = ext.all_emails[0]
    elif ext.persons:
        for p in ext.persons:
            if p.is_constituent and p.emails:
                ext.best_email = p.emails[0]
                break
        if not ext.best_email:
            for p in ext.persons:
                if p.emails:
                    ext.best_email = p.emails[0]
                    break

    # --- PHONE ---
    valid = [(s, p) for s, p in ext.all_phones if s == 'valid']
    if ext.persons:
        for p in ext.persons:
            if p.is_constituent and p.phones:
                vp = [(s, n) for s, n in p.phones if s == 'valid']
                if vp:
                    ext.best_phone = vp[0][1]
                    break
    if not ext.best_phone and valid:
        ext.best_phone = valid[0][1]
    if not ext.best_phone:
        partial = [(s, p) for s, p in ext.all_phones if s == 'partial']
        if partial:
            ext.best_phone = partial[0][1]

    # --- ADDRESS ---
    # Priority: parsed > constituent context > person block > single extracted
    if ext.parsed_street:
        ext.best_address = ext.parsed_street
    else:
        const_addrs = [a for a in ext.all_addresses if is_constituent_context(text, a)]
        if const_addrs:
            ext.best_address = const_addrs[0]
        elif ext.persons:
            for p in ext.persons:
                if p.is_constituent and p.addresses:
                    ext.best_address = p.addresses[0]
                    break
        if not ext.best_address and len(ext.all_addresses) == 1:
            ext.best_address = ext.all_addresses[0]

    # --- POSTAL ---
    if ext.parsed_postal:
        valid, norm = normalize_postal(ext.parsed_postal)
        if valid:
            ext.best_postal = norm
    if not ext.best_postal:
        valid_postals = [p for v, p in ext.all_postals if v]
        if valid_postals:
            ext.best_postal = valid_postals[0]

    # --- NAME ---
    if ext.persons:
        for p in ext.persons:
            if p.is_constituent and p.name:
                ext.best_name = p.name
                break
    if not ext.best_name and ext.all_names:
        ext.best_name = ext.all_names[0][0]


def _match_to_contacts(ext: CaseExtraction, ci: ContactIndex):
    """Try to match unlinked case to existing contact."""
    # Try email first
    emails_to_try = [ext.best_email] + ext.all_emails
    for email in emails_to_try:
        if not email:
            continue
        result = ci.match_email(email)
        if result:
            ext.matched_contact_id, ext.matched_contact_name, ext.match_method = result
            ext.flags.append(f"MATCHED:{ext.match_method}")
            return

    # Try phone
    if ext.best_phone:
        result = ci.match_phone(ext.best_phone)
        if result:
            ext.matched_contact_id, ext.matched_contact_name, ext.match_method = result
            ext.flags.append(f"MATCHED:{ext.match_method}")
            return


def _compute_fields(ext: CaseExtraction):
    """Compute missing and extractable fields."""
    if not ext.sf_contact_name:
        ext.missing_fields.append("name")
    if not ext.sf_contact_email or ext.sf_is_contaminated:
        ext.missing_fields.append("email")
    if not ext.sf_phone:
        ext.missing_fields.append("phone")

    if ext.best_email and (not ext.sf_contact_email or ext.sf_is_contaminated):
        ext.extractable_fields.append("email")
    if ext.best_phone and not ext.sf_phone:
        ext.extractable_fields.append("phone")
    if ext.best_address:
        ext.extractable_fields.append("address")
    if ext.best_postal:
        ext.extractable_fields.append("postal")
    if ext.best_name and not ext.sf_contact_name:
        ext.extractable_fields.append("name")


def _determine_action(ext: CaseExtraction) -> Tuple[str, str]:
    """Determine recommended action and confidence."""

    if ext.sf_is_contaminated:
        if ext.best_email and not is_staff_email(ext.best_email):
            return "FIX_CONTAMINATION_RECOVERABLE", "medium"
        return "FIX_CONTAMINATION_MANUAL", "low"

    if ext.sf_has_contact:
        if ext.extractable_fields:
            conf = "high" if not ext.is_email_chain else "medium"
            return "CAN_ENRICH", conf
        return "COMPLETE", "high"

    # Unlinked cases
    if ext.matched_contact_id:
        return "LINK_TO_EXISTING", "high" if 'exact' in ext.match_method else "medium"

    if ext.best_email:
        if ext.is_multi_person:
            if ext.llm_used and ext.llm_confidence in ('high', 'medium'):
                return "CREATE_CONTACT", "medium"
            return "CREATE_CONTACT_REVIEW", "low"
        if ext.is_email_chain and len(ext.all_emails) > 1:
            if ext.llm_used:
                return "CREATE_CONTACT", "medium"
            return "CREATE_CONTACT_REVIEW", "low"
        return "CREATE_CONTACT", "medium" if len(ext.all_emails) == 1 else "low"

    if ext.best_phone:
        return "HAS_PHONE_ONLY", "low"

    if ext.best_address or ext.best_name:
        return "HAS_PARTIAL_INFO", "low"

    return "NO_CONTACT_INFO", "none"


# duplicate detection

def find_duplicates(results: List[CaseExtraction]) -> Dict[str, List[int]]:
    email_to_idxs = defaultdict(list)
    for i, ext in enumerate(results):
        if not ext.sf_has_contact and ext.best_email:
            email_to_idxs[ext.best_email.lower()].append(i)
    return {e: idxs for e, idxs in email_to_idxs.items() if len(idxs) > 1}


# excel output

ACTION_COLORS = {
    'COMPLETE': '90EE90',
    'CAN_ENRICH': 'FFFACD',
    'LINK_TO_EXISTING': 'ADD8E6',
    'CREATE_CONTACT': '87CEEB',
    'CREATE_CONTACT_REVIEW': 'B0E0E6',
    'FIX_CONTAMINATION_RECOVERABLE': 'FFB6C1',
    'FIX_CONTAMINATION_MANUAL': 'FF9999',
    'HAS_PHONE_ONLY': 'DDA0DD',
    'HAS_PARTIAL_INFO': 'E6D8E6',
    'NO_CONTACT_INFO': 'D3D3D3',
    'NO_TEXT': 'C0C0C0',
}

ACTION_DESCRIPTIONS = {
    'COMPLETE': 'Has contact linked, no missing info found',
    'CAN_ENRICH': 'Contact linked but we found additional phone/address/name in case text',
    'LINK_TO_EXISTING': 'No contact linked but email/phone matches an existing contact — just link it',
    'CREATE_CONTACT': 'No contact linked, good email found — create new contact',
    'CREATE_CONTACT_REVIEW': 'No contact linked, email found but ambiguous (multi-person or chain) — review first',
    'FIX_CONTAMINATION_RECOVERABLE': 'Contact email is staff email, but real constituent email found in text',
    'FIX_CONTAMINATION_MANUAL': 'Contact email is staff email, no clear replacement found',
    'HAS_PHONE_ONLY': 'Phone number found but no email',
    'HAS_PARTIAL_INFO': 'Address or name found but no email or phone',
    'NO_CONTACT_INFO': 'Has text but no contact info could be extracted',
    'NO_TEXT': 'No text fields to extract from',
}

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def build_output_df(cases_df: pd.DataFrame, results: List[CaseExtraction],
                    duplicates: Dict[str, List[int]]) -> pd.DataFrame:
    """Build the output dataframe with all extraction columns."""

    # Mark duplicates
    for email, idxs in duplicates.items():
        for i in idxs:
            results[i].flags.append(f"DUP_GROUP:{email[:25]}({len(idxs)})")

    out = cases_df.copy()

    out['Action'] = [r.action for r in results]
    out['Confidence'] = [r.confidence for r in results]
    out['Missing_Fields'] = [', '.join(r.missing_fields) or '—' for r in results]
    out['Extractable_Fields'] = [', '.join(r.extractable_fields) or '—' for r in results]

    out['Best_Email'] = [r.best_email for r in results]
    out['Best_Phone'] = [r.best_phone for r in results]
    out['Best_Address'] = [r.best_address for r in results]
    out['Best_Postal'] = [r.best_postal for r in results]
    out['Best_Name'] = [r.best_name for r in results]

    out['SF_Has_Name'] = ['Yes' if r.sf_contact_name else 'No' for r in results]
    out['SF_Has_Email'] = ['Yes' if r.sf_contact_email and not r.sf_is_contaminated else 'No' for r in results]
    out['SF_Has_Phone'] = ['Yes' if r.sf_phone else 'No' for r in results]
    out['SF_Contaminated'] = ['Yes' if r.sf_is_contaminated else 'No' for r in results]

    out['Parsed_Address'] = [r.parsed_street for r in results]
    out['Parsed_Postal'] = [r.parsed_postal for r in results]

    out['All_Emails'] = ['; '.join(r.all_emails) for r in results]
    out['All_Phones'] = ['; '.join(f"{s}:{p}" for s, p in r.all_phones) for r in results]
    out['All_Addresses'] = ['; '.join(r.all_addresses) for r in results]
    out['All_Postals'] = ['; '.join(f"{'✓' if v else '✗'}{p}" for v, p in r.all_postals) for r in results]
    out['All_Names'] = ['; '.join(f"{n} ({t})" for n, t in r.all_names) for r in results]

    out['Person_Count'] = [r.person_count for r in results]
    out['Persons'] = [
        ' | '.join(
            f"{p.name or '?'}: {', '.join(p.emails)} / {', '.join(n for _,n in p.phones)} / {', '.join(p.addresses)} [{p.source}]"
            for p in r.persons
        ) for r in results
    ]

    out['LLM_Email'] = [r.llm_email for r in results]
    out['LLM_Phone'] = [r.llm_phone for r in results]
    out['LLM_Address'] = [r.llm_address for r in results]
    out['LLM_Complaint_Addr'] = [r.llm_complaint_addr for r in results]
    out['LLM_Confidence'] = [r.llm_confidence for r in results]
    out['LLM_Reasoning'] = [r.llm_reasoning for r in results]

    out['Matched_Contact_ID'] = [r.matched_contact_id for r in results]
    out['Matched_Contact_Name'] = [r.matched_contact_name for r in results]
    out['Match_Method'] = [r.match_method for r in results]

    out['Flags'] = ['; '.join(r.flags) for r in results]

    return out


def write_sheet(wb, name, df, max_rows=None):
    """Write a dataframe to a worksheet with formatting."""
    ws = wb.create_sheet(name)
    rows = list(dataframe_to_rows(df.head(max_rows) if max_rows else df, index=False, header=True))

    for r_idx, row_data in enumerate(rows, 1):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == 1:
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill('solid', fgColor='4472C4')
                cell.font = Font(bold=True, size=10, color='FFFFFF')
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Auto-width (capped)
    for c_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(c_idx)
        max_len = max(
            len(str(ws.cell(row=1, column=c_idx).value or '')),
            *(len(str(ws.cell(row=r, column=c_idx).value or '')[:50]) for r in range(2, min(len(rows) + 1, 20)))
        ) if len(rows) > 1 else 10
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Color rows by action if Action column exists
    if 'Action' in df.columns:
        action_col = list(df.columns).index('Action') + 1
        for r_idx in range(2, len(rows) + 1):
            action = ws.cell(row=r_idx, column=action_col).value
            if action in ACTION_COLORS:
                fill = PatternFill('solid', fgColor=ACTION_COLORS[action])
                for c_idx in range(1, min(len(df.columns) + 1, 50)):
                    ws.cell(row=r_idx, column=c_idx).fill = fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    return ws


def create_excel(output_df: pd.DataFrame, output_path: str, stats: dict,
                 duplicates: dict, results: List[CaseExtraction]):
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Dashboard"
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60

    r = 1
    ws.cell(row=r, column=1, value="CONSTITUENT EXTRACTOR v4 — RESULTS").font = Font(bold=True, size=16)
    r += 2

    # Stats section
    stat_items = [
        ("Total cases", stats['total']),
        ("Cases with Contact ID linked", stats['with_contact']),
        ("Cases without Contact ID", stats['without_contact']),
        ("Staff email contaminated", stats['contaminated']),
        ("Cases with no text at all", stats['no_text']),
        ("Multi-person cases detected", stats['multi_person']),
        ("LLM calls made", stats['llm_calls']),
    ]
    for label, val in stat_items:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="ACTIONS").font = Font(bold=True, size=14)
    ws.cell(row=r, column=2, value="Count").font = Font(bold=True)
    ws.cell(row=r, column=3, value="What To Do").font = Font(bold=True)
    r += 1

    for action in [
        'LINK_TO_EXISTING', 'FIX_CONTAMINATION_RECOVERABLE', 'CREATE_CONTACT',
        'CAN_ENRICH', 'CREATE_CONTACT_REVIEW', 'FIX_CONTAMINATION_MANUAL',
        'HAS_PHONE_ONLY', 'HAS_PARTIAL_INFO', 'NO_CONTACT_INFO', 'NO_TEXT', 'COMPLETE',
    ]:
        count = stats['actions'].get(action, 0)
        if count == 0:
            continue
        ws.cell(row=r, column=1, value=action)
        ws.cell(row=r, column=2, value=count)
        ws.cell(row=r, column=3, value=ACTION_DESCRIPTIONS.get(action, ''))
        if action in ACTION_COLORS:
            ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=ACTION_COLORS[action])
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="EXTRACTABLE FIELDS (can be added to Salesforce)").font = Font(bold=True, size=14)
    r += 1
    for fld, count in stats['extractable'].items():
        ws.cell(row=r, column=1, value=f"Could add {fld}")
        ws.cell(row=r, column=2, value=count)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="MATCHING").font = Font(bold=True, size=14)
    r += 1
    for method, count in stats.get('match_methods', {}).items():
        ws.cell(row=r, column=1, value=f"Matched via {method}")
        ws.cell(row=r, column=2, value=count)
        r += 1
    ws.cell(row=r, column=1, value="Duplicate email groups")
    ws.cell(row=r, column=2, value=stats['dup_groups'])
    r += 1
    ws.cell(row=r, column=1, value="Cases in duplicate groups")
    ws.cell(row=r, column=2, value=stats['dup_cases'])

    # --- Filtered action sheets (priority order) ---
    priority_sheets = [
        ('LINK_TO_EXISTING', 'Link'),
        ('FIX_CONTAMINATION_RECOVERABLE', 'Fix-Recover'),
        ('CREATE_CONTACT', 'Create'),
        ('CAN_ENRICH', 'Enrich'),
        ('CREATE_CONTACT_REVIEW', 'Create-Review'),
        ('FIX_CONTAMINATION_MANUAL', 'Fix-Manual'),
        ('HAS_PHONE_ONLY', 'Phone-Only'),
        ('HAS_PARTIAL_INFO', 'Partial'),
    ]

    # Slim column set for action sheets
    slim_cols = [
        'Case Number', 'Subject', 'Action', 'Confidence',
        'Best_Email', 'Best_Phone', 'Best_Address', 'Best_Postal', 'Best_Name',
        'Matched_Contact_ID', 'Matched_Contact_Name', 'Match_Method',
        'Person_Count', 'Persons',
        'Missing_Fields', 'Extractable_Fields',
        'Contact Name', 'Contact: Email', 'Web Email',
        'Flags',
    ]
    slim_cols = [c for c in slim_cols if c in output_df.columns]

    for action, sheet_name in priority_sheets:
        sub = output_df[output_df['Action'] == action]
        if len(sub) > 0:
            write_sheet(wb, sheet_name, sub[slim_cols])

    # --- Multi-person sheet ---
    multi = output_df[output_df['Person_Count'] > 1]
    if len(multi) > 0:
        multi_cols = slim_cols + ['Description']
        multi_cols = [c for c in multi_cols if c in output_df.columns]
        write_sheet(wb, 'Multi-Person', multi[multi_cols])

    # --- Duplicates sheet ---
    if duplicates:
        ws_dup = wb.create_sheet("Duplicates")
        ws_dup['A1'] = "Email"
        ws_dup['B1'] = "Count"
        ws_dup['C1'] = "Case Numbers"
        for cell in [ws_dup['A1'], ws_dup['B1'], ws_dup['C1']]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4472C4')
        row = 2
        for email, idxs in sorted(duplicates.items(), key=lambda x: -len(x[1])):
            cnums = [results[i].case_number for i in idxs]
            ws_dup.cell(row=row, column=1, value=email)
            ws_dup.cell(row=row, column=2, value=len(idxs))
            ws_dup.cell(row=row, column=3, value=', '.join(cnums))
            row += 1
        ws_dup.column_dimensions['A'].width = 40
        ws_dup.column_dimensions['C'].width = 80

    # --- All Cases sheet (last, since it's large) ---
    write_sheet(wb, 'All Cases', output_df)

    wb.save(output_path)


# main
def process(cases_path: str, contacts_path: Optional[str], output_path: str,
            use_llm: bool = False, llm_limit: int = 1500, api_key: str = "") -> dict:

    print(f"Loading cases from {cases_path}...")
    cases_df = load_file(cases_path)
    print(f"  {len(cases_df)} cases loaded")

    contacts_df = None
    if contacts_path:
        print(f"Loading contacts from {contacts_path}...")
        contacts_df = load_file(contacts_path)
        print(f"  {len(contacts_df)} contacts loaded")

    # LLM client
    client = None
    if use_llm and api_key and HAS_ANTHROPIC:
        client = anthropic.Anthropic(api_key=api_key)
        print("LLM disambiguation enabled")
    else:
        use_llm = False
        print("LLM disambiguation disabled (regex-only mode)")

    # Build contact index
    print("Building contact index...")
    ci = ContactIndex(contacts_df, cases_df)
    print(f"  {len(ci.email_to_contact)} unique emails, {len(ci.phone_to_contact)} phones indexed")

    # Process cases — two-pass approach
    # Pass 1: Extract everything with regex (fast, free)
    print("Pass 1: Regex extraction...")
    results: List[CaseExtraction] = []

    for idx, row in cases_df.iterrows():
        if idx % 1000 == 0 and idx > 0:
            print(f"  {idx}/{len(cases_df)} processed")
        ext = extract_case(row, ci, client=None, use_llm=False)
        results.append(ext)

    print(f"  {len(results)} cases extracted")

    # Pass 2: LLM disambiguation on highest-priority cases
    llm_calls = 0
    if use_llm and client:
        # Score each case for LLM priority
        llm_candidates = []
        for i, ext in enumerate(results):
            if not needs_llm(ext):
                continue
            score = _llm_priority_score(ext)
            llm_candidates.append((score, i, ext))

        # Sort by priority (highest first)
        llm_candidates.sort(key=lambda x: -x[0])
        eligible = len(llm_candidates)
        to_process = eligible if llm_limit == 0 else min(eligible, llm_limit)

        print(f"Pass 2: LLM disambiguation ({to_process} of {eligible} eligible, limit={llm_limit})...")
        if eligible > 0:
            # Show priority breakdown
            reasons = defaultdict(int)
            for score, _, ext in llm_candidates[:to_process]:
                if ext.is_multi_person:
                    reasons['multi_person'] += 1
                elif len(ext.all_emails) > 1:
                    reasons['multi_email'] += 1
                elif ext.is_email_chain:
                    reasons['email_chain'] += 1
                elif len(ext.all_addresses) > 1:
                    reasons['multi_address'] += 1
            print(f"  Priority breakdown: {dict(reasons)}")

        for rank, (score, idx, ext) in enumerate(llm_candidates[:to_process]):
            if rank % 100 == 0 and rank > 0:
                print(f"  {rank}/{to_process} LLM calls made")

            combined = ext._combined_text
            result = call_llm(combined, ext.all_emails, ext.all_phones,
                              ext.all_addresses, ext.persons, client)
            if result and 'error' not in result:
                ext.llm_used = True
                ext.llm_email = result.get('constituent_email') or ''
                ext.llm_phone = result.get('constituent_phone') or ''
                ext.llm_address = result.get('constituent_address') or ''
                ext.llm_complaint_addr = result.get('complaint_address') or ''
                ext.llm_confidence = result.get('confidence') or ''
                ext.llm_reasoning = result.get('reasoning') or ''
                ext.flags.append("LLM_USED")

                # LLM overrides best guess if it has an answer
                if ext.llm_email and not is_staff_email(ext.llm_email):
                    ext.best_email = ext.llm_email
                if ext.llm_phone:
                    ext.best_phone = ext.llm_phone
                if ext.llm_address:
                    ext.best_address = ext.llm_address

                # Re-match and re-classify after LLM update
                if not ext.sf_has_contact:
                    _match_to_contacts(ext, ci)
                _compute_fields(ext)
                ext.action, ext.confidence = _determine_action(ext)

                llm_calls += 1
            time.sleep(0.1)

    print(f"  {llm_calls} LLM calls made")

    # Duplicates
    duplicates = find_duplicates(results)
    print(f"  {len(duplicates)} duplicate email groups")

    # Build output
    output_df = build_output_df(cases_df, results, duplicates)

    # Stats
    match_methods = defaultdict(int)
    for r in results:
        if r.match_method:
            base = r.match_method.split('(')[0]
            match_methods[base] += 1

    stats = {
        'total': len(results),
        'with_contact': sum(1 for r in results if r.sf_has_contact),
        'without_contact': sum(1 for r in results if not r.sf_has_contact),
        'contaminated': sum(1 for r in results if r.sf_is_contaminated),
        'no_text': sum(1 for r in results if r.action == 'NO_TEXT'),
        'multi_person': sum(1 for r in results if r.is_multi_person),
        'llm_calls': llm_calls,
        'actions': output_df['Action'].value_counts().to_dict(),
        'extractable': {
            'email': sum(1 for r in results if 'email' in r.extractable_fields),
            'phone': sum(1 for r in results if 'phone' in r.extractable_fields),
            'address': sum(1 for r in results if 'address' in r.extractable_fields),
            'postal': sum(1 for r in results if 'postal' in r.extractable_fields),
            'name': sum(1 for r in results if 'name' in r.extractable_fields),
        },
        'match_methods': dict(match_methods),
        'dup_groups': len(duplicates),
        'dup_cases': sum(len(v) for v in duplicates.values()),
    }

    print(f"Writing to {output_path}...")
    create_excel(output_df, output_path, stats, duplicates, results)

    # Print summary
    print("\n" + "=" * 70)
    print("EXTRACTION COMPLETE")
    print("=" * 70)
    print(f"\nTotal: {stats['total']} cases")
    print(f"  Linked: {stats['with_contact']}  |  Unlinked: {stats['without_contact']}  |  Contaminated: {stats['contaminated']}")
    print(f"  Multi-person: {stats['multi_person']}  |  No text: {stats['no_text']}")

    print(f"\n--- ACTIONS ---")
    for action in [
        'LINK_TO_EXISTING', 'FIX_CONTAMINATION_RECOVERABLE', 'CREATE_CONTACT',
        'CAN_ENRICH', 'CREATE_CONTACT_REVIEW', 'FIX_CONTAMINATION_MANUAL',
        'HAS_PHONE_ONLY', 'HAS_PARTIAL_INFO', 'NO_CONTACT_INFO', 'NO_TEXT', 'COMPLETE',
    ]:
        count = stats['actions'].get(action, 0)
        if count:
            print(f"  {action:40s} {count:>5}")

    print(f"\n--- EXTRACTABLE ---")
    for fld, count in stats['extractable'].items():
        if count:
            print(f"  {fld:40s} {count:>5}")

    print(f"\n--- MATCHING ---")
    for method, count in stats['match_methods'].items():
        print(f"  {method:40s} {count:>5}")
    print(f"  {'Duplicate groups':40s} {stats['dup_groups']:>5}")

    print(f"\nOutput: {output_path}")
    return stats


def main():
    parser = argparse.ArgumentParser(description='Constituent Contact Extractor v4')
    parser.add_argument('cases', help='Cases export (CSV or XLSX)')
    parser.add_argument('contacts', nargs='?', help='Contacts export (CSV or XLSX)')
    parser.add_argument('-o', '--output', default='extraction_results.xlsx', help='Output file')
    parser.add_argument('--llm', action='store_true', help='Enable LLM disambiguation')
    parser.add_argument('--llm-limit', type=int, default=1500, help='Max LLM calls (default 1500, use 0 for unlimited)')
    parser.add_argument('--api-key', help='Anthropic API key')

    args = parser.parse_args()
    api_key = args.api_key or os.environ.get('ANTHROPIC_API_KEY', '')

    process(
        cases_path=args.cases,
        contacts_path=args.contacts,
        output_path=args.output,
        use_llm=args.llm,
        llm_limit=args.llm_limit,
        api_key=api_key,
    )


if __name__ == '__main__':
    main()
